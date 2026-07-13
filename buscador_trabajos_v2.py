"""
Buscador de trabajos - VERSION 3 -> Excel
=========================================
Cambios vs version anterior:
  * Computrabajo: selector de ubicacion corregido (antes capturaba el
    rating de la empresa tipo "4,2" en vez de la localidad).
  * Filtro por zona: CABA y/o GBA (AMBA). Descarta Rosario, Cordoba, etc.
  * Links CLICKEABLES en el Excel (hipervinculo real, no texto).
  * Indeed y Bumeran siguen incluidos, pero pueden traer 0 (ver notas
    al pie del archivo).

USO:
    python buscador_trabajos_v3.py
"""

import datetime
import json
import os
import re
import time
from pathlib import Path
import requests

# SSL: en esta maquina, Norton Antivirus intercepta HTTPS ("SSL/TLS
# scanning") y re-firma los certificados con su propia CA root. Windows
# confia en esa root (por eso Playwright/Bumeran no tiene problema), pero
# el bundle propio de certifi que usa `requests` no la incluye -> Computrabajo
# fallaba con CERTIFICATE_VERIFY_FAILED. `truststore` hace que el ssl
# de Python use el almacen de certificados del sistema operativo (el mismo
# que ya usa el navegador), en vez del bundle de certifi. Sigue verificando
# de verdad: no es un verify=False. Si el paquete no esta instalado, sigue
# con certifi/default (mismo comportamiento de antes, no rompe el script).
try:
    import truststore
    truststore.inject_into_ssl()
    print("SSL: usando truststore del sistema operativo")
except ImportError:
    print("SSL: truststore no disponible; requests usará certifi")

import pandas as pd
from openpyxl.styles import Font, PatternFill
from bs4 import BeautifulSoup

# ----------------------------------------------------------------------
# MODO DE USO / INTERRUPTORES
# ----------------------------------------------------------------------
# Para el uso de todos los dias, tocar SOLO MODO_EJECUCION:
#   "AUTO"   -> Programador de tareas: sin ventanas, no se cuelga esperando
#               un click.
#   "MANUAL" -> lo corres vos a mano y queres ver el resultado.
#   "TEST"   -> probar mejoras (Excel/fichas/resumen) aunque no haya
#               ofertas nuevas.
#   "DEMO"   -> genera un Excel de ejemplo con ofertas simuladas, SIN
#               internet y SIN tocar el historial. Sirve para mostrar el
#               proyecto o probar el formato sin depender de Bumeran/
#               Computrabajo/Indeed.
MODO_EJECUCION = "MANUAL"  # "AUTO", "MANUAL", "TEST" o "DEMO"

# --- independientes de MODO_EJECUCION: se dejan como estaban ------------
# True = activa Indeed.
# False = deja Indeed apagado. Recomendado porque suele bloquear con Cloudflare.
ACTIVAR_INDEED = False

# True = guarda screenshot/html si Indeed es bloqueado.
# False = no guarda archivos debug.
GUARDAR_DEBUG_INDEED = True

# True = descarta avisos de Bumeran que redirigen a ZonaJobs u otros externos.
# False = permite todos los avisos de Bumeran.
DESCARTAR_BUMERAN_EXTERNOS = True

# Maximo de dias de antiguedad de los avisos.
# Mantener en 2 para buscar solo avisos recientes.
CT_DIAS = 2

# Sonido al terminar: "OFF" (nada), "SUAVE" (corto y tranquilo) o
# "NORMAL" (el de siempre, mas fuerte).
SONIDO_FINAL = "SUAVE"  # "OFF", "SUAVE" o "NORMAL"

# --- estas 5 se configuran automaticamente segun MODO_EJECUCION ---------
# ANALIZAR_DESCRIPCION_DETALLE: entra al detalle de cada aviso nuevo
# POSTULAR/REVISAR (con el mismo navegador de Bumeran/Indeed) para leer
# la descripcion completa, armar un resumen, detectar alertas (senioridad,
# ingles alto, monotributo, posible estafa, etc.) y ajustar decision/
# prioridad segun eso. Es lo que mas tiempo agrega, por eso va apagado en
# AUTO (Programador de tareas) y prendido en MANUAL/TEST.
NAVEGADOR_VISIBLE = False
MOSTRAR_CARTEL_FINAL = False
ABRIR_EXCEL_AL_FINAL = False
MODO_TEST_GENERAR_EXCEL_AUNQUE_SEAN_REPETIDAS = False
ANALIZAR_DESCRIPCION_DETALLE = False

if MODO_EJECUCION == "AUTO":
    # Programador de tareas: sin navegador visible, sin messagebox (nadie
    # para clickear "Aceptar"), sin abrir el Excel solo, y sin entrar al
    # detalle de cada aviso (es lo que mas tarda) para correr rapido.
    NAVEGADOR_VISIBLE = False
    MOSTRAR_CARTEL_FINAL = False
    ABRIR_EXCEL_AL_FINAL = False
    MODO_TEST_GENERAR_EXCEL_AUNQUE_SEAN_REPETIDAS = False
    ANALIZAR_DESCRIPCION_DETALLE = False

elif MODO_EJECUCION == "MANUAL":
    NAVEGADOR_VISIBLE = False
    MOSTRAR_CARTEL_FINAL = True
    ABRIR_EXCEL_AL_FINAL = True
    MODO_TEST_GENERAR_EXCEL_AUNQUE_SEAN_REPETIDAS = False
    ANALIZAR_DESCRIPCION_DETALLE = True

elif MODO_EJECUCION == "TEST":
    NAVEGADOR_VISIBLE = False
    MOSTRAR_CARTEL_FINAL = True
    ABRIR_EXCEL_AL_FINAL = True
    MODO_TEST_GENERAR_EXCEL_AUNQUE_SEAN_REPETIDAS = True
    ANALIZAR_DESCRIPCION_DETALLE = True

elif MODO_EJECUCION == "DEMO":
    # Sin internet: no llama a Bumeran/Computrabajo/Indeed, no toca
    # vistos.json ni historial_trabajos.xlsx. Usa ofertas simuladas
    # (generar_datos_demo) para mostrar el Excel/RESUMEN completo.
    NAVEGADOR_VISIBLE = False
    MOSTRAR_CARTEL_FINAL = True
    ABRIR_EXCEL_AL_FINAL = True
    MODO_TEST_GENERAR_EXCEL_AUNQUE_SEAN_REPETIDAS = False
    ANALIZAR_DESCRIPCION_DETALLE = False

else:
    raise ValueError('MODO_EJECUCION debe ser "AUTO", "MANUAL", "TEST" o "DEMO"')


# ----------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------
BUSQUEDAS = [
    # --- compuestas (precisas) ---
    "anotacion de datos", "data annotation", "data annotator", "AI trainer",
    "AI data trainer", "machine learning data", "evaluador de IA", "prompt engineer",
    "prompt evaluator",
    "analista de datos", "data analyst", "junior data analyst", "analista bi",
    "analista sql", "analista de datos junior", "data analyst junior",
    "SQL junior", "Power BI junior", "data entry remote",
    "desarrollador junior", "junior developer", "programador junior",
    "desarrollador web", "desarrollador python", "python developer",
    "QA tester", "QA tester junior", "QA tester remote", "analista de testing",
    "soporte IT", "soporte IT junior", "soporte tecnico", "mesa de ayuda",
    "help desk", "analista funcional", "analista funcional junior",
    "trainee IT", "automation trainee", "administrativo sistemas",
    "back office sistemas",
    "analista supply chain", "planificador de produccion",
    "analista de abastecimiento", "analista de compras",
    "analista de inventarios", "administrativo logistica", "asistente de logistica",
    # --- palabras solas (mas volumen; el filtro fino contiene el ruido) ---
    "analista",
    "desarrollador", "programador", "soporte", "administrativo", "logistica",
    "abastecimiento", "compras",
    "inventario", "datos", "testing", "qa", "sistemas",
    # --- planning / IT ---
    "planner", "planning", "planeamiento", "planificacion",
    "analista IT", "reclutador IT", "perfil IT",
    # --- sueltas extra (ES) + sus equivalentes en ingles ---
    "tester",
    "developer", "programmer", "software developer", "web developer",
    "backend developer", "frontend developer", "full stack",
    "support", "IT support", "helpdesk", "service desk",
    "QA analyst", "quality assurance", "software tester",
    "data entry", "data engineer",
    "supply chain", "supply chain analyst", "warehouse", "logistics",
    "planner demand",
]

PERFIL_KEYWORDS = [
    "ia", "ai", "llm", "anotacion", "annotation", "prompt", "datos", "data",
    "python", "sql", "power bi", "bi", "excel", "oracle", "analista",
    "machine learning", "junior", "trainee", "intern", "entry level",
    "desarrollador", "programador", "developer", "web", "html", "css",
    "javascript", "qa", "testing", "soporte", "erp", "remote", "remoto",
    "annotator", "evaluator", "trainer", "helpdesk", "help desk", "back office",
    "automation", "supply", "logistica", "inventario",
    "abastecimiento", "compras", "planificacion", "almacen", "administrativo",
    "tecnico",
]

# --- FILTRO DE ZONA ---------------------------------------------------
# Solo se guardan ofertas cuya ubicacion contenga alguna de estas
# palabras. Para no filtrar nada, deja la lista vacia: ZONAS_OK = []
ZONAS_OK = [
    "capital federal", "caba", "ciudad autonoma",   # CABA
    "buenos aires-gba", "gba", "buenos aires",       # AMBA / provincia
]
# Si una oferta NO tiene ubicacion detectable, ¿la guardo igual?
# True = no perder avisos remotos o sin dato. False = solo zona confirmada.
GUARDAR_SIN_UBICACION = True

UBICACION_INDEED = "Buenos Aires"
MAX_POR_FUENTE = 20

# --- HISTORIAL / MEMORIA -----------------------------------------------
# El link es la clave: si ya esta en vistos.json, el aviso no se vuelve a
# mostrar en el Excel del dia (aunque siga apareciendo en las busquedas).
HISTORIAL_JSON = "vistos.json"
HISTORIAL_XLSX = "historial_trabajos.xlsx"

# --- POSTULACION ASISTIDA/AUTOMATICA (segura) ---------------------------
# "OFF"         -> no hace nada, comportamiento actual (default).
# "ASISTIDA"    -> abre cada oferta POSTULAR HOY y llega hasta el
#                  formulario, pero SIEMPRE se detiene antes de enviar.
#                  La decision de mandarla la toma Sergio, a mano.
# "AUTO_SEGURO" -> evalua condiciones de seguridad (sin alertas graves,
#                  sin redirect externo, sin CAPTCHA, dominio propio de
#                  Bumeran/Computrabajo). Con DRY_RUN_POSTULACION=True
#                  solo simula (no toca nada real). Ver postular_oferta()
#                  para la nota sobre por que, aunque exista este modo,
#                  el envio real automatico desatendido no esta
#                  implementado todavia.
# "CONFIRMADA"  -> pregunta por consola, aviso por aviso, antes de
#                  intentar enviar. SOLO funciona con MODO_EJECUCION=
#                  "MANUAL" (alguien tiene que estar presente para
#                  contestar el prompt); en AUTO/TEST/DEMO se desactiva
#                  solo, para no dejar el script colgado esperando una
#                  respuesta que nunca va a llegar. Ver postular_oferta()
#                  para la nota sobre el click real todavia sin conectar.
MODO_POSTULACION = "CONFIRMADA"  # "OFF", "ASISTIDA", "AUTO_SEGURO" o "CONFIRMADA"

# True = ninguna postulacion real se envia, todo queda simulado/registrado.
# Dejar en True hasta confiar en el criterio del agente.
DRY_RUN_POSTULACION = True

# Techo de seguridad: cuantas ofertas como maximo se procesan por corrida.
MAX_POSTULACIONES_POR_CORRIDA = 1

# False (default) = en MODO_EJECUCION="TEST" la postulacion no se procesa,
# igual que ahora (solo se genera el Excel con las repetidas).
# True = en TEST tambien se corre el flujo de postulacion (apertura,
# chequeos de seguridad, y confirmacion si corresponde) sobre las ofertas
# de ACCIONES -- para poder probarlo sin esperar ofertas nuevas reales.
# Sigue respetando MAX_POSTULACIONES_POR_CORRIDA y DRY_RUN_POSTULACION, y
# NUNCA toca vistos.json ni historial_trabajos.xlsx (eso ya esta separado
# de esta funcion). Cada resultado queda marcado como prueba de TEST.
PROBAR_POSTULACION_EN_TEST = False

if MODO_POSTULACION not in ("OFF", "ASISTIDA", "AUTO_SEGURO", "CONFIRMADA"):
    raise ValueError('MODO_POSTULACION debe ser "OFF", "ASISTIDA", "AUTO_SEGURO" o "CONFIRMADA"')

# Log de cada intento de postulacion (uno por fila procesada), separado
# del historial de avisos vistos.
POSTULACIONES_LOG_XLSX = "postulaciones_log.xlsx"

# Carpeta donde se guardan los Excel del dia (normales y de TEST),
# organizados en subcarpetas por mes: resultados/YYYY-MM/. vistos.json,
# historial_trabajos.xlsx, perfil_sergio.txt y log_buscador.txt NO se
# mueven: se quedan en la carpeta principal del script.
RESULTADOS_DIR = "resultados"

# --- LOCK (evita 2 instancias corriendo a la vez) -----------------------
# Protege vistos.json, historial_trabajos.xlsx y los Excel de que dos
# corridas (ej. una manual desde VS Code + la tarea programada) escriban
# al mismo tiempo. "No iniciar una instancia nueva" del Programador de
# tareas solo controla instancias lanzadas por LA TAREA, no una corrida
# manual, por eso hace falta este lock propio.
LOCK_FILE = "buscador.lock"
# Si el lock es mas viejo que esto, se asume abandonado (crash, Task
# Manager, corte de luz) y se ignora en vez de bloquear para siempre.
LOCK_MAX_HORAS = 3

# --- PERFIL DEL USUARIO --------------------------------------------------
# Archivo de texto con el resumen del perfil de Sergio. Si no existe, se
# crea solo con un contenido base la primera vez que corre el script.
PERFIL_SERGIO_PATH = "perfil_sergio.txt"

# Texto del perfil, cargado por cargar_perfil_usuario() al arrancar main().
# Placeholder vacio por si algo llegara a leerse antes de esa carga.
PERFIL_USUARIO_TEXTO = ""

# --- DECISION SUGERIDA ---------------------------------------------------
# Capa de clasificacion sobre lo que ya paso los filtros. NO elimina nada
# (ni siquiera los DESCARTAR quedan afuera del Excel): solo marca, para
# que la decision final la siga tomando Sergio. Trabaja sobre lo que hay
# disponible en ese punto (titulo, empresa, ubicacion, modalidad, busqueda),
# no sobre la descripcion completa del aviso (esa ya se descarto antes,
# solo se uso para calcular 'relevancia').
DECISION_POSTULAR_KEYWORDS = [
    "analista de datos", "data analyst", "sql", "power bi", "excel", "python",
    "soporte it", "mesa de ayuda", "help desk", "analista funcional", "qa", "testing",
    "desarrollador junior", "programador junior", "trainee", "junior",
    "ai", "annotation", "annotator", "trainer", "prompt", "back office", "automation",
    "remoto", "híbrido", "hibrido", "caba", "capital federal", "supply chain",
    "abastecimiento", "compras", "inventario", "planificación", "planificacion",
    "administrativo",
]

# Si aparece cualquiera de estas, DESCARTAR gana sobre POSTULAR (se chequea
# primero). Las frases largas ("mas de 3 años excluyente", etc.) solo
# matchean si literalmente aparecen en el titulo -- no hay descripcion
# completa disponible en este punto para buscarlas con mas contexto.
DECISION_DESCARTAR_KEYWORDS = [
    "senior", "ssr fuerte", "más de 3 años excluyente", "mas de 3 años excluyente",
    "experiencia excluyente", "ventas puras", "ventas puro", "call center",
    "gastronomía", "gastronomia", "vendedor", "cajero", "rrhh", "reclutador",
    "comercial puro", "presencial lejos", "zonajobs", "link externo",
    "inglés c1 excluyente", "ingles c1 excluyente",
]

# Umbral: con esta cantidad o mas de keywords de POSTULAR matcheadas, se
# marca POSTULAR. Menos que eso pero al menos 1 match -> REVISAR.
DECISION_POSTULAR_MIN_MATCHES = 2

# --- Filtros que Computrabajo aplica EN LA URL (mas confiable) ---
# Ciudad: tal como aparece en la URL de Computrabajo. Ej: "capital-federal".
# Vacio ("") = todo el pais.
CT_CIUDAD = "capital-federal"
# Paginas a recorrer por busqueda. El scraper avanza p=1, p=2, ... y se
# DETIENE solo cuando una pagina viene vacia. CT_MAX_PAGINAS es solo un
# tope de seguridad para no quedar en bucle infinito si el sitio falla.
CT_MAX_PAGINAS = 20
# Pausa (segundos) entre cada pedido, para no parecer bot y evitar el 403.
CT_PAUSA = 1.0

# --- FILTRO FINO (se aplica despues de traer los avisos) -------------
# 1) Relevancia minima: descarta avisos con menos de N keywords de tu
#    perfil. Sube esto para mas precision (menos ruido, menos volumen).
#    0 = no filtra. 2 = recomendado para empezar.
MIN_RELEVANCIA = 2
# 2) Match de titulo: solo conserva avisos cuyo TITULO contenga alguna
#    de estas palabras. Asi se van cocineros, pizzeros, cajeros, etc.
#    Lista vacia = no filtra por titulo.
TITULO_DEBE_CONTENER = [
    # data / IA / IT
    "analista", "datos", "data", "python", "sql", "bi", "desarrollador",
    "programador", "developer", "qa", "tester", "testing", "soporte",
    "it", "sistemas", "funcional", "tecnico", "técnico",
    "planner", "planning",
    # AI / prompt / trainee (antes faltaban -> avisos tipo "AI Trainer" se
    # perdian en este filtro aunque estuvieran en BUSQUEDAS y PERFIL_KEYWORDS)
    "ai", "ia", "trainer", "prompt", "annotation", "annotator", "evaluator",
    "junior", "trainee", "helpdesk",
    # supply chain / admin
    "supply", "abastecimiento", "compras", "inventario", "logistica",
    "logística", "planificador", "planeamiento", "administrativo",
    "almacen", "almacén", "deposito", "depósito",
]
# Palabras cortas/ambiguas que deben matchear como PALABRA ENTERA, no como
# fragmento (si no, "it" matchea "digital", "bi" matchea "habil", etc.)
TITULO_PALABRA_ENTERA = {"it", "bi", "qa", "sql"}

# LISTA NEGRA: si el titulo contiene alguna de estas, se DESCARTA aunque
# haya pasado el filtro de arriba. Mata el ruido que se cuela por palabras
# ambiguas (ej. "Cocinero con exp en produccion" entra por "produccion").
TITULO_NO_DEBE_CONTENER = [
    # gastronomia
    "cocin", "gastro", "chef", "parrill", "mozo", "moza", "camarer",
    "bachero", "pizzer", "sushi", "panader", "pasteler", "barista",
    "ayudante de cocina",
    # ventas / atencion / otros rubros lejanos a tu perfil
    "vendedor", "vendedora", "cajero", "cajera", "promotor", "ventas",
    "telemarket", "call center", "atencion al cliente",
    # RRHH / reclutamiento
    "recursos humanos", "rrhh", "reclut", "selección de personal",
    "seleccion de personal", "capital humano", "talent", "generalista de rrhh",
    "capacitación y desarrollo", "capacitacion y desarrollo",
    # industria / fabrica / mecanica (fuera de perfil IT/Data/Soporte)
    "mantenimiento", "operario", "producción", "produccion",
    "mecánico", "mecanico", "electromecánico", "electromecanico",
    "técnico industrial", "tecnico industrial",
]

# --- INDEED ------------------------------------------------------------
# Indeed puede bloquear con una pantalla de verificacion de Cloudflare
# (no es un problema de selectores: la pagina de resultados real ni
# siquiera llega a cargar). Se activa/desactiva con ACTIVAR_INDEED, en
# el panel de interruptores arriba del archivo.

# Textos que aparecen cuando Indeed muestra la pantalla de Cloudflare/
# CAPTCHA en vez de los resultados reales.
BLOQUEOS_INDEED = [
    "cloudflare",
    "verificación adicional requerida",
    "verificacion adicional requerida",
    "tu id de ray",
    "ray id",
    "captcha",
    "robot",
    "verificando",
    "unusual traffic",
    "access denied",
]

# --- FILTRO DE POSTULACION EXTERNA (Bumeran) --------------------------
# Bumeran a veces solo replica un aviso publicado en otro portal (ZonaJobs,
# LinkedIn, etc.) y el boton "Postularme" te saca de Bumeran. Se
# activa/desactiva con DESCARTAR_BUMERAN_EXTERNOS, en el panel de
# interruptores arriba del archivo.

BUMERAN_EXTERNOS_BLOQUEADOS = [
    "zonajobs",
    "postularme en zonajobs",
    "el botón te dirigirá al aviso completo",
    "link directo",
    "tinyurl",
    "bit.ly",
    "hiringroom",
    "linkedin",
    "indeed",
    "google forms",
]

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
COLUMNAS = ["relevancia", "titulo", "empresa", "ubicacion", "modalidad",
            "fecha", "link", "busqueda", "fuente"]

# patron que detecta un rating mal capturado como ubicacion: "4", "4,2", "3.5"
_RATING = re.compile(r"^\s*\d([.,]\d)?\s*$")


def _ubicacion_valida(texto):
    """Devuelve la ubicacion si parece real, o None si es un rating/basura."""
    if not texto:
        return None
    t = texto.strip()
    if _RATING.match(t):       # es un rating tipo "4,2"
        return None
    if len(t) < 3:
        return None
    return t


def _pasa_zona(ubic):
    if not ZONAS_OK:
        return True
    if not ubic:
        return GUARDAR_SIN_UBICACION
    u = ubic.lower()
    return any(z in u for z in ZONAS_OK)


# ----------------------------------------------------------------------
# Fuentes
# ----------------------------------------------------------------------

# --- Navegador Playwright compartido (se abre 1 vez, lo reusan Bumeran
#     e Indeed; abrir uno por busqueda seria lentisimo). NAVEGADOR_VISIBLE
#     se controla desde el panel de interruptores arriba del archivo. ----
PW_PAUSA = 1.5             # segundos de espera tras cargar cada pagina
DEBUG_SCRAPER = True       # logs de diagnostico por fuente (url, titulo pagina,
                           # cantidad de links/tarjetas, bloqueo detectado)

_PW = {"play": None, "browser": None, "page": None}

# Cache de verificacion de postulacion externa en Bumeran: href -> (es_externo, motivo).
# Muchas BUSQUEDAS se solapan (ej. "analista" y "analista de datos" traen el
# mismo aviso), asi que sin esto se re-verificaria el mismo link muchas veces.
_BUMERAN_VERIFICADOS = {}


def _get_page():
    """Devuelve una pagina de navegador lista, abriendo el navegador la
    primera vez. Reutiliza la misma en llamadas siguientes."""
    if _PW["page"] is not None:
        return _PW["page"]
    print(f"Navegador visible: {NAVEGADOR_VISIBLE}")
    from playwright.sync_api import sync_playwright
    _PW["play"] = sync_playwright().start()
    _PW["browser"] = _PW["play"].chromium.launch(
        headless=not NAVEGADOR_VISIBLE
    )
    ctx = _PW["browser"].new_context(
        user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/131.0.0.0 Safari/537.36"),
        locale="es-AR",
    )
    _PW["page"] = ctx.new_page()
    return _PW["page"]


def _cerrar_navegador():
    try:
        if _PW["browser"]:
            _PW["browser"].close()
        if _PW["play"]:
            _PW["play"].stop()
    except Exception:
        pass
    _PW["page"] = _PW["browser"] = _PW["play"] = None


# Selectores candidatos para la tarjeta de aviso en Indeed, en orden de
# preferencia (Indeed cambia el HTML seguido).
_INDEED_SELECTORES_TARJETA = [
    "div.job_seen_beacon",
    "div.cardOutline",
    "td.resultContent",
    "div.slider_container",
]
_INDEED_SELECTORES_TITULO = ["h2.jobTitle", "h2 a span", "a.jcs-JobTitle"]
_INDEED_SELECTORES_EMPRESA = ["[data-testid='company-name']", "span.companyName"]
_INDEED_SELECTORES_UBICACION = ["[data-testid='text-location']", "div.companyLocation"]
_INDEED_SELECTORES_FECHA = ["[data-testid='myJobsStateDate']", "span.date"]

# Contador global: cuantas veces Indeed mostro Cloudflare/CAPTCHA en esta
# corrida. main() lo usa para aclarar en el diagnostico final que el 0 de
# Indeed es por bloqueo, no por selectores rotos.
_INDEED_BLOQUEOS_DETECTADOS = 0


def _indeed_primer_match(elemento, selectores):
    """Prueba selectores en orden y devuelve el texto del primero que
    matchee (o "" si ninguno matchea)."""
    for sel in selectores:
        try:
            tag = elemento.query_selector(sel)
        except Exception:
            tag = None
        if tag:
            texto = tag.inner_text().strip()
            if texto:
                return texto
    return ""


def indeed(query):
    """Indeed Argentina via navegador real (Playwright)."""
    global _INDEED_BLOQUEOS_DETECTADOS
    filas = []
    try:
        page = _get_page()
        from urllib.parse import quote_plus
        url = (f"https://ar.indeed.com/jobs?q={quote_plus(query)}"
               f"&l={quote_plus(UBICACION_INDEED)}&fromage={CT_DIAS}")

        print(f"    [Indeed] URL usada: {url}")
        print(f"    [Indeed] CT_DIAS aplicado: {CT_DIAS}")

        page.goto(url, timeout=45000, wait_until="domcontentloaded")
        try:
            page.wait_for_selector(_INDEED_SELECTORES_TARJETA[0], timeout=8000)
        except Exception:
            pass
        time.sleep(PW_PAUSA)

        cuerpo = (page.inner_text("body") or "")

        tarjetas = []
        for sel in _INDEED_SELECTORES_TARJETA:
            try:
                candidatas = page.query_selector_all(sel)
            except Exception:
                candidatas = []
            if candidatas:
                tarjetas = candidatas
                break

        print(f"    [Indeed] tarjetas encontradas: {len(tarjetas)}")

        if not tarjetas:
            # El chequeo de bloqueo se hace SOLO si no hay tarjetas: con
            # tarjetas reales en pantalla, palabras sueltas de BLOQUEOS_INDEED
            # como "robot" o "verificando" pueden venir de un aviso normal
            # (ej. "analista de robots RPA", "verificando indicadores") y no
            # significan Cloudflare. Una pagina de bloqueo real, en cambio,
            # nunca tiene tarjetas.
            cuerpo_low = cuerpo.lower()
            bloqueo = [kw for kw in BLOQUEOS_INDEED if kw in cuerpo_low]
            if bloqueo:
                _INDEED_BLOQUEOS_DETECTADOS += 1
                print(f"    [Indeed] Cloudflare/CAPTCHA detectado ({', '.join(bloqueo)}) "
                      "-> se salta esta búsqueda")
                if GUARDAR_DEBUG_INDEED:
                    try:
                        page.screenshot(path="debug_indeed_cloudflare.png", full_page=True)
                        html = page.content()
                        with open("debug_indeed_cloudflare.html", "w", encoding="utf-8") as f:
                            f.write(html)
                        print("    [Indeed] evidencia guardada: "
                              "debug_indeed_cloudflare.png / .html")
                    except Exception as e:
                        print(f"    [Indeed] no se pudo guardar evidencia de debug: {e}")
            elif DEBUG_SCRAPER:
                print(f"    [Indeed] 0 tarjetas. Primeros textos en pantalla: {cuerpo[:200]!r}")
            return filas

        # Indeed a veces repite una tarjeta (mismo titulo+empresa) con un
        # link todavia sin hidratar del lado del cliente; se dedupea por
        # contenido ademas del dedup global que ya hace main() con el Excel.
        vistos = set()
        for t in tarjetas[:MAX_POR_FUENTE]:
            try:
                titulo = _indeed_primer_match(t, _INDEED_SELECTORES_TITULO)
                if not titulo:
                    continue
                empresa = _indeed_primer_match(t, _INDEED_SELECTORES_EMPRESA)
                clave = (titulo.lower(), empresa.lower())
                if clave in vistos:
                    continue
                vistos.add(clave)
                ubicacion = _ubicacion_valida(_indeed_primer_match(t, _INDEED_SELECTORES_UBICACION))
                fecha = _indeed_primer_match(t, _INDEED_SELECTORES_FECHA)

                a = t.query_selector("h2.jobTitle a") or t.query_selector("a.jcs-JobTitle")
                href = a.get_attribute("href") if a else ""
                link = ("https://ar.indeed.com" + href) if href and href.startswith("/") else (href or "")

                filas.append({
                    "titulo": titulo, "empresa": empresa,
                    "ubicacion": ubicacion,
                    "modalidad": "", "fecha": fecha,
                    "link": link, "busqueda": query, "fuente": "Indeed",
                    "_texto": f"{titulo} {empresa} {ubicacion or ''}",
                })
            except Exception:
                continue
        time.sleep(PW_PAUSA)
    except Exception as e:
        print(f"  [Indeed] error: {e}")
    return filas


# Selectores candidatos para la tarjeta de aviso en Bumeran, en orden de
# preferencia. Bumeran cambia clases hasheadas (styled-components) seguido;
# "a[href*='/empleos/']" es el mas estable porque depende de la URL, no de
# una clase. Si el sitio cambia de nuevo, agregar mas opciones ac
# ("por que" del orden: se prueba cada uno y se usa el primero que matchee).
_BUMERAN_SELECTORES_TARJETA = [
    "a[href*='/empleos/']",
    "article a[href*='empleo']",
    "div[id^='listado'] a[href]",
]

# Etiquetas de modalidad conocidas, para clasificar los <h3> de cada tarjeta.
_BUMERAN_MODALIDADES = {"presencial", "remoto", "hibrido", "híbrido", "mixto"}


def _clasificar_h3_bumeran(textos):
    """Los <h3> de una tarjeta de Bumeran traen, en orden variable:
    fecha, empresa, (a veces) rating de empresa, ubicacion, modalidad.
    Los clasifica por contenido en vez de por posicion, porque el orden
    y la cantidad cambian segun si el aviso tiene rating o no."""
    fecha = empresa = ubicacion = modalidad = ""
    resto = []
    for t in textos:
        tl = (t or "").strip()
        if not tl:
            continue
        low = tl.lower()
        if low.startswith("actualizado") or low.startswith("publicado"):
            fecha = tl
        elif low in _BUMERAN_MODALIDADES:
            modalidad = tl
        elif _RATING.match(tl):
            continue  # rating de empresa tipo "3.7", no sirve
        elif "," in tl:
            ubicacion = _ubicacion_valida(tl)
        else:
            resto.append(tl)
    if resto and not empresa:
        empresa = resto[0]
    return fecha, empresa, ubicacion, modalidad


def _chequear_postulacion_bumeran(page, url):
    """Visita el detalle de un aviso de Bumeran y dice si la postulacion
    es propia de Bumeran o te manda a un portal externo (ZonaJobs, etc).
    Devuelve (es_externo, motivo):
      True, "zonajobs"  -> aviso puente, hay que descartarlo.
      False, None       -> postulacion dentro de Bumeran, conservar.
      None, None        -> no se pudo cargar el detalle; se conserva el
                            aviso igual para no perder resultados validos
                            por un error de red puntual."""
    try:
        page.goto(url, timeout=20000, wait_until="domcontentloaded")
        # el detalle hidrata con React: recien hay contenido real cuando
        # aparece el <h1> del titulo. Sin esto, a veces se lee la pagina
        # "vacia" (solo nav/footer) y el chequeo da un falso negativo.
        try:
            page.wait_for_selector("h1", timeout=8000)
        except Exception:
            pass
        cuerpo = (page.inner_text("body") or "").lower()
    except Exception:
        return None, None
    for kw in BUMERAN_EXTERNOS_BLOQUEADOS:
        if kw in cuerpo:
            return True, kw
    return False, None


_BUMERAN_BLOQUEO_KW = ("captcha", "cloudflare", "robot", "acceso denegado",
                       "verifica que sos humano", "unusual traffic")


def _bumeran_cargar_listado(page, url):
    """Abre una URL de listado de Bumeran y devuelve (tarjetas, titulo_pagina,
    total_links, cuerpo_texto). No lanza excepcion: si algo falla devuelve
    listas/strings vacias para que el llamador decida que hacer."""
    try:
        page.goto(url, timeout=45000, wait_until="domcontentloaded")
        # esperar carga real: probar el selector principal en vez de un
        # sleep fijo a ciegas; si no aparece en el tiempo dado, seguimos
        # igual (puede ser que la busqueda de 0 resultados reales).
        try:
            page.wait_for_selector(_BUMERAN_SELECTORES_TARJETA[0], timeout=8000)
        except Exception:
            pass
        time.sleep(PW_PAUSA)

        titulo_pagina = page.title()
        total_links = len(page.query_selector_all("a"))

        tarjetas = []
        for sel in _BUMERAN_SELECTORES_TARJETA:
            try:
                candidatas = page.query_selector_all(sel)
            except Exception:
                candidatas = []
            if candidatas:
                tarjetas = candidatas
                break

        cuerpo = page.inner_text("body") or ""
        return tarjetas, titulo_pagina, total_links, cuerpo
    except Exception as e:
        print(f"    [Bumeran] error cargando {url}: {e}")
        return [], "", 0, ""


def _bumeran_detectar_bloqueo(cuerpo):
    cuerpo_low = cuerpo.lower()
    return [kw for kw in _BUMERAN_BLOQUEO_KW if kw in cuerpo_low]


def _coincide_query_local(texto, query):
    """Usado solo en el fallback zona+fecha (sin keyword en la URL): decide
    si una tarjeta trae la busqueda pedida, mirando titulo+empresa+
    descripcion completa de la tarjeta. Exige que TODAS las palabras
    relevantes (>=3 letras) de la query esten en el texto, en cualquier
    orden, para tolerar que la redaccion del aviso no sea idéntica."""
    texto_low = texto.lower()
    q = query.lower().strip()
    if q in texto_low:
        return True
    palabras = [w for w in re.split(r"\s+", q) if len(w) >= 3]
    if not palabras:
        return False
    return all(w in texto_low for w in palabras)


def _bumeran_extraer_candidatos(tarjetas, query, filtrar_localmente):
    """PASADA 1: lee todo lo necesario de cada tarjeta de la pagina de
    listado, sin navegar todavia (si navegamos antes de terminar de leer
    todas las tarjetas, Playwright invalida los ElementHandle viejos)."""
    candidatos = []
    vistos = set()
    for t in tarjetas[:MAX_POR_FUENTE * 2]:
        try:
            href = t.get_attribute("href") or ""
            if "/empleos/" not in href or href in vistos:
                continue
            vistos.add(href)

            texto_completo = t.inner_text()

            h2 = t.query_selector("h2")
            titulo = h2.inner_text().strip() if h2 else ""
            if not titulo:
                # fallback defensivo si el sitio deja de usar <h2>:
                # la primera linea que no sea fecha ni rating.
                for linea in texto_completo.split("\n"):
                    linea = linea.strip()
                    if (linea and not linea.lower().startswith(("actualizado", "publicado"))
                            and not _RATING.match(linea)):
                        titulo = linea
                        break
            if not titulo or len(titulo) < 4:
                continue

            if filtrar_localmente and not _coincide_query_local(texto_completo, query):
                continue

            h3s = [h.inner_text() for h in t.query_selector_all("h3")]
            fecha, empresa, ubicacion, modalidad = _clasificar_h3_bumeran(h3s)

            link = href if href.startswith("http") else ("https://www.bumeran.com.ar" + href)
            candidatos.append({
                "titulo": titulo, "empresa": empresa,
                "ubicacion": ubicacion,
                "modalidad": modalidad, "fecha": fecha,
                "link": link, "busqueda": query, "fuente": "Bumeran",
                "_texto": f"{titulo} {empresa}",
            })
        except Exception:
            continue
    return candidatos


def bumeran(query):
    """Bumeran Argentina via navegador real (Playwright).

    CT_DIAS se aplica asi:
      1) Se intenta la URL combinada palabra clave + fecha
         (empleos-publicacion-menor-a-N-dias-busqueda-{slug}.html).
      2) Bumeran no siempre tiene ese bucket para busquedas por palabra
         clave (los cortes reales suelen arrancar en 3 dias, no en 1-2).
         Si la combinada da 0 tarjetas, se cae a la pagina de zona+fecha
         (Buenos Aires, sin keyword) y se filtra localmente por la query
         usando el texto de cada tarjeta (titulo+empresa+descripcion).
    """
    filas = []
    try:
        page = _get_page()
        from urllib.parse import quote
        slug = quote(query.strip().lower().replace(" ", "-"))

        if CT_DIAS:
            url = (f"https://www.bumeran.com.ar/empleos-publicacion-menor-a-"
                   f"{CT_DIAS}-dias-busqueda-{slug}.html")
        else:
            url = f"https://www.bumeran.com.ar/empleos-busqueda-{slug}.html"

        print(f"    [Bumeran] URL usada: {url}")
        print(f"    [Bumeran] CT_DIAS aplicado: {CT_DIAS}")

        tarjetas, titulo_pagina, total_links, cuerpo = _bumeran_cargar_listado(page, url)

        bloqueo = _bumeran_detectar_bloqueo(cuerpo)
        if bloqueo:
            print(f"    [Bumeran] BLOQUEO detectado ({', '.join(bloqueo)}) -> no se pudo leer")
            return filas

        filtrar_localmente = False
        if not tarjetas and CT_DIAS:
            url_fallback = (f"https://www.bumeran.com.ar/en-buenos-aires/"
                             f"empleos-publicacion-menor-a-{CT_DIAS}-dias.html")
            print(f"    [Bumeran] combinado dio 0 tarjetas -> fallback zona+fecha: {url_fallback}")
            tarjetas, titulo_pagina, total_links, cuerpo = _bumeran_cargar_listado(page, url_fallback)
            filtrar_localmente = True

            bloqueo = _bumeran_detectar_bloqueo(cuerpo)
            if bloqueo:
                print(f"    [Bumeran] BLOQUEO detectado ({', '.join(bloqueo)}) -> no se pudo leer")
                return filas

        if DEBUG_SCRAPER:
            print(f"    [Bumeran] titulo pagina: {titulo_pagina!r}")
            print(f"    [Bumeran] <a> totales: {total_links} | tarjetas: {len(tarjetas)}")

        if not tarjetas:
            if DEBUG_SCRAPER:
                print(f"    [Bumeran] 0 tarjetas. Primeros textos en pantalla: "
                      f"{cuerpo[:200]!r}")
            return filas

        # --- PASADA 1: extraer candidatos de la pagina de listado ---------
        candidatos = _bumeran_extraer_candidatos(tarjetas, query, filtrar_localmente)

        # --- PASADA 2: verificar si la postulacion es propia de Bumeran ---
        # Se visita el detalle de cada candidato (con cache global para no
        # re-chequear el mismo aviso si aparece en otra busqueda).
        for fila in candidatos:
            if len(filas) >= MAX_POR_FUENTE:
                break

            if not DESCARTAR_BUMERAN_EXTERNOS:
                filas.append(fila)
                continue

            link = fila["link"]
            veredicto = _BUMERAN_VERIFICADOS.get(link)
            if veredicto is None:
                veredicto = _chequear_postulacion_bumeran(page, link)
                _BUMERAN_VERIFICADOS[link] = veredicto
            es_externo, motivo = veredicto

            if es_externo is True:
                print(f"    [Bumeran] descartado por postulación externa: {motivo} "
                      f"({fila['titulo']})")
                continue
            if es_externo is None:
                print(f"    [Bumeran] no se pudo verificar el detalle, se conserva "
                      f"igual: {fila['titulo']}")
            elif DEBUG_SCRAPER:
                print(f"    [Bumeran] guardado: postulación Bumeran ({fila['titulo']})")

            filas.append(fila)

        if DEBUG_SCRAPER:
            print(f"    [Bumeran] filas guardadas: {len(filas)} de {len(candidatos)} candidatos")
        time.sleep(PW_PAUSA)
    except Exception as e:
        print(f"  [Bumeran] error: {e}")
    return filas


def _parse_tarjeta_ct(art, query):
    """Extrae una oferta de una tarjeta <article> de Computrabajo."""
    titulo_tag = art.select_one("a.js-o-link")
    if not titulo_tag:
        return None
    titulo = titulo_tag.get_text(strip=True)

    # Ubicacion: el HTML real usa <span class="mr10">Localidad</span>.
    ubic = None
    for sel in ["span.mr10", "p.fs16.fc_base.mt5", "p.fc_base.mt5"]:
        tag = art.select_one(sel)
        if tag:
            cand = _ubicacion_valida(tag.get_text(" ", strip=True))
            if cand:
                ubic = cand
                break
    if not ubic:
        txt = art.get_text(" ", strip=True)
        m = re.search(r"([A-ZÁÉÍÓÚ][\w áéíóú]+,\s*(Capital Federal|"
                      r"Buenos Aires[\w\- ]*|Santa Fe|Córdoba))", txt)
        if m:
            ubic = m.group(1).strip()

    empresa_tag = art.select_one("a.fc_base") or art.select_one("p.dFlex a")

    fecha_txt = ""
    for tag in art.find_all(["p", "span"]):
        t = tag.get_text(" ", strip=True)
        if t.lower().startswith("hace ") or t.lower() in ("ayer", "hoy"):
            fecha_txt = t
            break

    return {
        "titulo": titulo,
        "empresa": empresa_tag.get_text(strip=True) if empresa_tag else "",
        "ubicacion": ubic,
        "modalidad": "", "fecha": fecha_txt,
        "link": "https://ar.computrabajo.com" + titulo_tag.get("href", ""),
        "busqueda": query, "fuente": "Computrabajo",
        "_texto": f"{titulo} {art.get_text(' ', strip=True)}",
    }


def computrabajo(query):
    filas = []
    ciudad = f"-en-{CT_CIUDAD}" if CT_CIUDAD else ""
    url = f"https://ar.computrabajo.com/trabajo-de-{query.strip().lower().replace(' ', '-')}{ciudad}"

    vistos_pagina = set()  # para detectar si una pagina repite la anterior
    for pagina in range(1, CT_MAX_PAGINAS + 1):
        try:
            params = {}
            if CT_DIAS:
                params["pubdate"] = CT_DIAS
            if pagina > 1:
                params["p"] = pagina
            r = requests.get(url, params=params, headers=HEADERS, timeout=20)
            if r.status_code == 403:
                print(f"    pagina {pagina}: 403 (bloqueo) -> corto esta busqueda")
                break
            soup = BeautifulSoup(r.text, "html.parser")
            tarjetas = soup.select("article.box_offer")
            if not tarjetas:
                break  # no hay mas avisos: fin real de la paginacion

            # firma de la pagina para no repetir (Computrabajo a veces
            # devuelve la ultima pagina existente si pedis una de mas)
            firma = tuple(t.select_one("a.js-o-link").get_text(strip=True)
                          for t in tarjetas if t.select_one("a.js-o-link"))
            if firma in vistos_pagina:
                break  # pagina repetida: ya recorrimos todo
            vistos_pagina.add(firma)

            for art in tarjetas:
                fila = _parse_tarjeta_ct(art, query)
                if fila:
                    filas.append(fila)

            time.sleep(CT_PAUSA)  # pausa para no gatillar el 403
        except Exception as e:
            print(f"    pagina {pagina}: error {e}")
            break
    return filas


# Bumeran + Computrabajo activas siempre. Indeed queda opcional (puede
# bloquear con Cloudflare) y se prende/apaga solo con ACTIVAR_INDEED de
# mas arriba en la config.
FUENTES = [bumeran, computrabajo]

if ACTIVAR_INDEED:
    FUENTES.insert(0, indeed)


def puntuar(fila):
    texto = (fila.get("_texto") or "").lower()
    return sum(1 for kw in PERFIL_KEYWORDS if kw.lower() in texto)


def adquirir_lock():
    """Crea LOCK_FILE si no existe. Si ya existe pero es mas viejo que
    LOCK_MAX_HORAS, se asume que quedo de una corrida que se corto mal
    (crash, Task Manager, corte de luz) y se pisa en vez de bloquear el
    buscador para siempre. Devuelve True si se puede correr, False si
    hay otra instancia corriendo de verdad ahora mismo."""
    if os.path.exists(LOCK_FILE):
        edad_horas = (time.time() - os.path.getmtime(LOCK_FILE)) / 3600
        if edad_horas < LOCK_MAX_HORAS:
            print("El buscador ya se está ejecutando. Se cancela esta instancia.")
            return False
        print(f"  Aviso: había un {LOCK_FILE} de hace {edad_horas:.1f}h "
              "(corrida anterior que se cortó mal). Lo ignoro y sigo.")

    try:
        with open(LOCK_FILE, "w", encoding="utf-8") as f:
            f.write(f"pid={os.getpid()}\n")
            f.write(f"inicio={datetime.datetime.now().isoformat()}\n")
    except Exception as e:
        print(f"  Aviso: no se pudo crear {LOCK_FILE}: {e}. Sigo sin lock.")
    return True


def liberar_lock():
    try:
        if os.path.exists(LOCK_FILE):
            os.remove(LOCK_FILE)
    except Exception as e:
        print(f"No se pudo borrar el lock: {e}")


def construir_ruta_excel(es_test=False, es_demo=False):
    """Arma la ruta resultados/YYYY-MM/trabajos_[TEST_|DEMO_]timestamp.xlsx
    y crea la carpeta (y la del mes) si no existen. vistos.json,
    historial_trabajos.xlsx, perfil_sergio.txt y log_buscador.txt NO
    pasan por aca: se siguen quedando en la carpeta principal."""
    # con segundos: evita colision si el lock permite dos corridas muy
    # seguidas o si se corre el script dos veces en el mismo minuto.
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    carpeta_mes = datetime.datetime.now().strftime("%Y-%m")
    base = Path(RESULTADOS_DIR) / carpeta_mes
    base.mkdir(parents=True, exist_ok=True)

    if es_demo:
        prefijo = "trabajos_DEMO"
    elif es_test:
        prefijo = "trabajos_TEST"
    else:
        prefijo = "trabajos"
    return base / f"{prefijo}_{timestamp}.xlsx"


def abrir_excel_si_corresponde(archivo):
    """Abre el Excel del dia con la app default de Windows. Solo se llama
    cuando el archivo realmente se genero (nunca en los caminos donde el
    script corta sin resultados). Si falla, no rompe el script."""
    if not ABRIR_EXCEL_AL_FINAL:
        return
    try:
        os.startfile(str(archivo))
    except Exception as e:
        print(f"No se pudo abrir el Excel automáticamente: {e}")


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def _formatear_duracion(inicio):
    """Devuelve la duracion desde 'inicio' hasta ahora, como texto
    'Xm Ys' (o 'Hh Xm Ys' si pasa la hora)."""
    segundos_totales = int((datetime.datetime.now() - inicio).total_seconds())
    horas, resto = divmod(segundos_totales, 3600)
    minutos, segundos = divmod(resto, 60)
    if horas:
        return f"{horas}h {minutos}m {segundos}s"
    return f"{minutos}m {segundos}s"


def avisar_fin(mensaje="El buscador terminó."):
    """Aviso de que el script termino: sonido + ventana emergente + log en
    consola. El script puede tardar 10-20+ min, asi que esto avisa sin
    tener que quedarse mirando la consola. Si algo del aviso falla (sin
    sonido, sin entorno grafico, etc.), nunca debe cortar el programa.

    El messagebox se controla con MOSTRAR_CARTEL_FINAL: en el Programador
    de tareas no hay nadie para clickear "Aceptar", asi que con False no
    se abre ninguna ventana y la tarea no queda colgada. El mensaje en
    consola se imprime siempre, para que quede en el log si la tarea
    redirige la salida a un archivo (ej. log_buscador.txt).

    El sonido se controla con SONIDO_FINAL: OFF no suena, SUAVE es un
    beep corto, NORMAL es el sonido de siempre."""
    if SONIDO_FINAL != "OFF":
        try:
            import winsound
            if SONIDO_FINAL == "SUAVE":
                winsound.Beep(700, 180)
            else:  # "NORMAL"
                winsound.MessageBeep()
                winsound.Beep(1000, 700)
        except Exception:
            pass

    if MOSTRAR_CARTEL_FINAL:
        try:
            import tkinter as tk
            from tkinter import messagebox
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            messagebox.showinfo("Buscador de trabajos", mensaje)
            root.destroy()
        except Exception as e:
            print(f"No se pudo mostrar el cartel emergente: {e}")

    print(f"\n{'=' * 60}")
    print(mensaje)
    print(f"{'=' * 60}\n")


def _cargar_vistos():
    """Lee vistos.json: {link: fecha_primera_vez}. Si no existe, esta
    vacio o esta corrupto, devuelve {} sin romper el script."""
    if not os.path.exists(HISTORIAL_JSON):
        return {}
    try:
        with open(HISTORIAL_JSON, "r", encoding="utf-8") as f:
            contenido = f.read().strip()
            if not contenido:
                return {}
            return json.loads(contenido)
    except Exception as e:
        print(f"  Aviso: no se pudo leer {HISTORIAL_JSON} ({e}), arranco de cero.")
        return {}


def _guardar_vistos(vistos):
    try:
        with open(HISTORIAL_JSON, "w", encoding="utf-8") as f:
            json.dump(vistos, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"  Aviso: no se pudo guardar {HISTORIAL_JSON}: {e}")


def _actualizar_historial_xlsx(df_nuevas, hoy):
    """Agrega las ofertas nuevas de hoy al historial acumulado. Si el
    archivo no existe o esta corrupto, lo crea de cero (no rompe)."""
    try:
        historial_previo = pd.read_excel(HISTORIAL_XLSX)
    except Exception:
        historial_previo = pd.DataFrame(columns=list(COLUMNAS) + ["primera_vez"])

    df_agregar = df_nuevas.copy()
    df_agregar["primera_vez"] = hoy

    try:
        historial_actualizado = pd.concat([historial_previo, df_agregar], ignore_index=True)
        historial_actualizado.to_excel(HISTORIAL_XLSX, index=False)
    except Exception as e:
        print(f"  Aviso: no se pudo actualizar {HISTORIAL_XLSX}: {e}")


POSTULACIONES_LOG_COLUMNAS = [
    "fecha_hora", "titulo", "empresa", "fuente", "link", "accion_sugerida",
    "estado_postulacion", "resultado_postulacion", "detalle_postulacion",
]


def _registrar_postulacion_log(registro):
    """Agrega UNA fila (dict) a postulaciones_log.xlsx. Append-only, mismo
    patron que _actualizar_historial_xlsx. Si falla, avisa y sigue: un
    problema con el log no debe frenar el resto de la corrida."""
    try:
        log_previo = pd.read_excel(POSTULACIONES_LOG_XLSX)
    except Exception:
        log_previo = pd.DataFrame(columns=POSTULACIONES_LOG_COLUMNAS)

    try:
        fila_nueva = pd.DataFrame([registro], columns=POSTULACIONES_LOG_COLUMNAS)
        log_actualizado = pd.concat([log_previo, fila_nueva], ignore_index=True)
        log_actualizado.to_excel(POSTULACIONES_LOG_XLSX, index=False)
    except Exception as e:
        print(f"  Aviso: no se pudo actualizar {POSTULACIONES_LOG_XLSX}: {e}")


_ORDEN_DECISION = {"POSTULAR": 0, "REVISAR": 1, "DESCARTAR": 2}


def clasificar_decision(fila):
    """Capa de decision sobre un aviso que ya paso todos los filtros.
    No descarta nada: solo sugiere. Devuelve (decision_sugerida,
    motivo_decision, prioridad)."""
    texto = " ".join(str(fila.get(c, "") or "") for c in
                      ("titulo", "empresa", "ubicacion", "modalidad", "busqueda")).lower()

    descartar_hits = [kw for kw in DECISION_DESCARTAR_KEYWORDS if kw in texto]
    if descartar_hits:
        return "DESCARTAR", f"excluido por: {', '.join(descartar_hits)}", 0

    postular_hits = [kw for kw in DECISION_POSTULAR_KEYWORDS if kw in texto]
    cant = len(postular_hits)

    if cant >= DECISION_POSTULAR_MIN_MATCHES:
        return "POSTULAR", f"match fuerte: {', '.join(postular_hits)}", cant
    if cant >= 1:
        return "REVISAR", f"match parcial: {', '.join(postular_hits)}", cant
    return "REVISAR", "sin señales claras en título/ubicación, revisar manualmente", 0


# --- FICHA DE POSTULACION (solo para POSTULAR) ---------------------------
# Todo por reglas simples, sin APIs ni modelos: se clasifica el aviso en
# una categoria segun que palabras de DECISION_POSTULAR_KEYWORDS matchean,
# y cada categoria tiene sus textos canonicos (porque conviene, como
# encaja el perfil, mensaje sugerido, accion recomendada).
_FICHA_CATEGORIAS = {
    "datos": {
        "keywords": ["analista de datos", "data analyst", "sql", "power bi", "python", "excel"],
        "porque_conviene": "Coincide con tu perfil de datos, SQL, Excel y análisis.",
        "encaje_perfil": ("Podés destacar SQL, Excel, Power BI y tu experiencia analizando "
                           "información operativa en Supply Chain."),
        "accion_recomendada": "Preparar mensaje orientado a datos.",
        "mensaje_sugerido": (
            "Hola, me interesa postularme a la posición. Soy estudiante de Ingeniería en "
            "Informática y cuento con experiencia en áreas operativas y de Supply Chain, "
            "trabajando con análisis de información, Excel, procesos y coordinación entre "
            "sectores. Además, estoy fortaleciendo mis conocimientos en SQL, Python y Power BI. "
            "Me gustaría poder aportar desde un perfil analítico, ordenado y con muchas ganas "
            "de seguir creciendo."
        ),
    },
    "soporte": {
        "keywords": ["soporte it", "mesa de ayuda", "qa", "testing", "analista funcional"],
        "porque_conviene": "Puede servir como entrada al área IT por soporte/mesa de ayuda.",
        "encaje_perfil": ("Podés orientar tu postulación hacia soporte, resolución de "
                           "problemas y conocimientos básicos de sistemas."),
        "accion_recomendada": "Preparar mensaje orientado a soporte IT.",
        "mensaje_sugerido": (
            "Hola, me interesa postularme a la posición. Soy estudiante de Ingeniería en "
            "Informática y vengo de roles operativos donde resolví problemas del día a día, "
            "coordiné con distintas áreas y usé sistemas de gestión y Excel de forma "
            "constante. Estoy sumando conocimientos técnicos en SQL, Python y bases de "
            "sistemas, y busco un rol donde pueda aportar capacidad de resolución, orden "
            "y ganas de aprender."
        ),
    },
    "supply_chain": {
        "keywords": ["supply chain", "abastecimiento", "compras", "inventario",
                     "planificación", "planificacion", "administrativo"],
        "porque_conviene": "Encaja con tu experiencia previa en Supply Chain, compras o planificación.",
        "encaje_perfil": ("Conviene remarcar tu experiencia en planificación, compras y "
                           "mejora de procesos."),
        "accion_recomendada": "Preparar mensaje orientado a Supply Chain.",
        "mensaje_sugerido": (
            "Hola, me interesa postularme a la posición. Cuento con experiencia previa en "
            "Supply Chain, participando en planificación, compras, abastecimiento e "
            "inventarios, y manejando Excel para el seguimiento de procesos. Actualmente "
            "estoy cursando Ingeniería en Informática, lo que me da una mirada analítica "
            "extra para este tipo de rol. Me gustaría sumar ese perfil ordenado y orientado "
            "a procesos a su equipo."
        ),
    },
    "junior_general": {
        "keywords": ["trainee", "junior", "desarrollador junior", "programador junior"],
        "porque_conviene": "Parece una oportunidad junior/trainee alineada a transición IT.",
        "encaje_perfil": ("Podés presentarte como perfil junior en transición hacia IT, "
                           "con base técnica en formación y experiencia laboral previa."),
        "accion_recomendada": "Postular hoy.",
        "mensaje_sugerido": (
            "Hola, me interesa postularme a la posición. Soy estudiante de Ingeniería en "
            "Informática, con conocimientos en Python, SQL, HTML, CSS y JavaScript, y con "
            "experiencia laboral previa en áreas operativas donde desarrollé orden, "
            "análisis y trabajo en equipo. Busco un primer paso dentro de IT donde pueda "
            "seguir formándome y aportar desde ya con compromiso y ganas de aprender."
        ),
    },
}

# Orden en que se evaluan las categorias cuando el aviso matchea keywords
# de mas de una (ej. "Analista de Datos Junior"): la mas especifica gana.
_FICHA_ORDEN_CATEGORIAS = ["datos", "supply_chain", "soporte", "junior_general"]

_PERFIL_SERGIO_CONTENIDO_BASE = """Perfil de ejemplo.
Estudiante de Ingeniería en Informática.
Experiencia previa en Supply Chain, planificación, compras, abastecimiento, inventarios, coordinación operativa y mejora de procesos.
Conocimientos técnicos: Python, SQL, Power BI, Excel, HTML, CSS, JavaScript.
Intereses laborales: Data, análisis de datos, soporte IT, mesa de ayuda, QA/testing, análisis funcional, desarrollo web junior, automatización, procesos, supply chain y roles administrativos vinculados a tecnología.
Idiomas: inglés intermedio e italiano básico.
Preferencia de comunicación: enfocar experiencia transferible, aprendizaje, orden, análisis y ganas de crecer.
"""


def cargar_perfil_usuario():
    """Lee PERFIL_SERGIO_PATH. Si no existe, lo crea con un contenido base
    la primera vez. Nunca rompe el script: si falla la lectura/escritura
    (permisos, disco, etc.), devuelve "" y el resto sigue funcionando con
    las categorias/plantillas fijas de siempre."""
    try:
        if not os.path.exists(PERFIL_SERGIO_PATH):
            with open(PERFIL_SERGIO_PATH, "w", encoding="utf-8") as f:
                f.write(_PERFIL_SERGIO_CONTENIDO_BASE)
            print(f"  Se creó {PERFIL_SERGIO_PATH} con el perfil base.")
        with open(PERFIL_SERGIO_PATH, "r", encoding="utf-8") as f:
            return f.read().strip()
    except Exception as e:
        print(f"  Aviso: no se pudo cargar {PERFIL_SERGIO_PATH}: {e}")
        return ""


def _extraer_linea_perfil(perfil_texto, prefijo):
    """Busca en el perfil una linea 'Prefijo: resto' y devuelve 'resto'.
    Si no la encuentra (o el perfil esta vacio), devuelve ""."""
    for linea in (perfil_texto or "").splitlines():
        linea = linea.strip()
        if linea.lower().startswith(prefijo.lower()):
            partes = linea.split(":", 1)
            return partes[1].strip().rstrip(".") if len(partes) > 1 else ""
    return ""


def generar_ficha_postulacion(fila):
    """Solo tiene sentido llamarla para avisos POSTULAR. Devuelve
    (porque_conviene, encaje_perfil, mensaje_sugerido, accion_recomendada),
    elegidos por reglas simples (que categoria de keywords matcheo mas),
    sin ninguna API ni modelo externo. 'encaje_perfil' se enriquece con
    lo que diga PERFIL_USUARIO_TEXTO (perfil_sergio.txt), asi que si se
    edita ese archivo, la ficha lo refleja sin tocar codigo."""
    texto = " ".join(str(fila.get(c, "") or "") for c in
                      ("titulo", "empresa", "ubicacion", "modalidad", "busqueda")).lower()

    mejor_categoria, mejor_cant = None, 0
    for nombre in _FICHA_ORDEN_CATEGORIAS:
        datos_cat = _FICHA_CATEGORIAS[nombre]
        cant = sum(1 for kw in datos_cat["keywords"] if kw in texto)
        if cant > mejor_cant:
            mejor_categoria, mejor_cant = nombre, cant

    if not mejor_categoria:
        mejor_categoria = "junior_general"

    cat = _FICHA_CATEGORIAS[mejor_categoria]

    encaje_perfil = cat["encaje_perfil"]
    conocimientos = _extraer_linea_perfil(PERFIL_USUARIO_TEXTO, "Conocimientos técnicos")
    if conocimientos:
        encaje_perfil = f"{encaje_perfil} (Perfil: {conocimientos})"

    return (cat["porque_conviene"], encaje_perfil,
            cat["mensaje_sugerido"], cat["accion_recomendada"])


# --- CATEGORIA DETECTADA --------------------------------------------------
# Etiqueta informativa (que tipo de puesto es), separada de la capa de
# decision_sugerida/prioridad -- no las toca ni las usa para decidir nada.
CATEGORIAS_KEYWORDS = {
    "Data / BI": ["data", "datos", "analista de datos", "data analyst", "sql",
                  "power bi", "bi", "reporting", "dashboard", "excel avanzado",
                  "python", "analytics"],
    "Soporte IT": ["soporte it", "soporte técnico", "soporte tecnico", "mesa de ayuda",
                   "help desk", "helpdesk", "service desk", "técnico en sistemas",
                   "tecnico en sistemas", "incidentes", "tickets", "redes",
                   "hardware", "software"],
    "QA / Testing": ["qa", "tester", "testing", "quality assurance",
                      "casos de prueba", "test cases", "automatización de pruebas",
                      "automatizacion de pruebas", "selenium", "jira"],
    "Desarrollo": ["desarrollador", "developer", "programador", "frontend",
                   "backend", "full stack", "javascript", "html", "css",
                   "react", "node", "python developer"],
    "Analista Funcional": ["analista funcional", "functional analyst", "requerimientos",
                            "user stories", "historias de usuario", "backlog",
                            "documentación funcional", "documentacion funcional",
                            "uat", "procesos"],
    "Supply Chain": ["supply chain", "abastecimiento", "compras", "inventario",
                      "logística", "logistica", "planificación", "planificacion",
                      "planeamiento", "planner", "almacén", "almacen",
                      "deposito", "depósito"],
    "Administrativo / Procesos": ["administrativo", "administración", "administracion",
                                    "asistente administrativo", "carga de datos",
                                    "excel", "procesos", "facturación", "facturacion",
                                    "documentación", "documentacion",
                                    "coordinación administrativa", "coordinacion administrativa"],
    "Técnico / Producción": ["técnico", "tecnico", "mecánico", "mecanico",
                              "mantenimiento", "producción", "produccion",
                              "operario", "electromecánico", "electromecanico",
                              "industrial"],
}

CATEGORIA_OTRO = "Otro"

# Orden de desempate cuando 2+ categorias matchean la misma cantidad de
# keywords: la primera de esta lista gana.
_ORDEN_CATEGORIAS_DETECCION = [
    "Data / BI", "Desarrollo", "QA / Testing", "Analista Funcional",
    "Soporte IT", "Supply Chain", "Administrativo / Procesos",
    "Técnico / Producción",
]


def detectar_categoria(fila):
    """Etiqueta 'que tipo de puesto es' en base a titulo/empresa/ubicacion/
    modalidad/busqueda + descripcion_resumen/alertas_aviso si existen
    (pueden venir vacios si ANALIZAR_DESCRIPCION_DETALLE=False). Gana la
    categoria con mas keywords matcheadas; empate lo resuelve
    _ORDEN_CATEGORIAS_DETECCION. Sin matches -> 'Otro'. Reglas simples,
    no toca decision_sugerida ni prioridad."""
    campos = ("titulo", "empresa", "ubicacion", "modalidad", "busqueda",
              "descripcion_resumen", "alertas_aviso")
    texto = " ".join(str(fila.get(c, "") or "") for c in campos).lower()

    conteos = {nombre: sum(1 for kw in kws if kw in texto)
               for nombre, kws in CATEGORIAS_KEYWORDS.items()}
    mejor_cant = max(conteos.values(), default=0)
    if mejor_cant == 0:
        return CATEGORIA_OTRO

    for nombre in _ORDEN_CATEGORIAS_DETECCION:
        if conteos.get(nombre, 0) == mejor_cant:
            return nombre
    return CATEGORIA_OTRO


# --- ANALISIS DE DESCRIPCION (detalle del aviso) --------------------------
# Todo por reglas simples (busqueda de texto / regex), nada de IA ni APIs.
# Ojo con las limitaciones honestas de este enfoque:
#   - "descripcion_resumen" es un recorte del texto, no un resumen real
#     generado (eso requeriria un modelo de lenguaje, que no se usa aca).
#   - "consultora sospechosa" no se puede detectar de forma confiable sin
#     una lista de empresas conocidas; en su lugar se buscan frases tipicas
#     de estafa ("pagar para postular", "curso pago", etc.).
#   - "salario muy bajo" es aproximado: busca el primer monto en pesos que
#     aparece en el texto y lo compara contra un umbral fijo. Si el aviso
#     no menciona salario, no genera alerta (no es sospechoso por omision).
#   - "tecnologias imposibles" no se intenta detectar (no hay forma
#     confiable con reglas simples); en su lugar se usa la combinacion de
#     "muchas señales de seniority" + "ningun skill del perfil mencionado"
#     como proxy de "este aviso no es para vos".

ALERTAS_AVISO_KEYWORDS = [
    "senior", "ssr", "semi senior", "semi-senior",
    "más de 3 años", "mas de 3 años", "4 años", "5 años", "6 años", "7 años",
    "inglés avanzado", "ingles avanzado", "inglés c1", "ingles c1",
    "excluyente", "monotributo", "contractor",
    "presencial lejos", "full presencial", "100% presencial",
    "zonajobs", "curso pago", "capacitación paga", "capacitacion paga",
    "pagar para postular", "inscripción con costo", "inscripcion con costo",
]

# Umbral (ARS) para la alerta de "salario bajo": solo se dispara si el
# aviso menciona un monto y ese monto esta por debajo de esto.
SALARIO_MINIMO_ALERTA = 700000

# Subconjunto de ALERTAS_AVISO_KEYWORDS que cuenta como "señal de
# seniority alta" para bajar decision/prioridad.
_ALERTAS_SENIORIDAD = [
    "senior", "ssr", "semi senior", "semi-senior",
    "más de 3 años", "mas de 3 años", "4 años", "5 años", "6 años", "7 años",
    "excluyente", "inglés avanzado", "ingles avanzado", "inglés c1", "ingles c1",
]
_SKILLS_POSITIVOS = ["sql", "excel", "power bi", "python", "soporte it",
                     "mesa de ayuda", "qa", "testing", "analisis funcional",
                     "análisis funcional", "supply chain"]
_SIGNALS_JUNIOR = ["junior", "trainee", "estudiante"]

_MONTO_PESOS_RE = re.compile(r"\$\s?(\d{1,3}(?:[.,]\d{3})+)")


def _resumir_texto(texto, largo=300):
    """'Resumen' simple sin IA: recorte del texto a un largo razonable,
    cortando en el ultimo espacio para no partir una palabra a la mitad."""
    texto = " ".join((texto or "").split())  # normaliza espacios/saltos
    if len(texto) <= largo:
        return texto
    corte = texto[:largo].rsplit(" ", 1)[0]
    return corte + "..."


def _detectar_alertas(texto):
    """Devuelve las alertas encontradas en el texto, como un solo string
    separado por '; ' (para que entre en una celda de Excel)."""
    texto_low = (texto or "").lower()
    encontradas = [kw for kw in ALERTAS_AVISO_KEYWORDS if kw in texto_low]

    m = _MONTO_PESOS_RE.search(texto or "")
    if m:
        monto_str = m.group(1).replace(".", "").replace(",", "")
        try:
            monto = int(monto_str)
            if monto < SALARIO_MINIMO_ALERTA:
                encontradas.append(f"salario bajo (~${monto:,})".replace(",", "."))
        except ValueError:
            pass

    return "; ".join(encontradas)


def extraer_detalle_aviso(fila):
    """Visita el link del aviso (Bumeran, Computrabajo o Indeed, no importa
    la fuente: se lee el texto de la pagina de forma generica) y devuelve
    (descripcion_texto, descripcion_resumen, alertas_aviso). Si algo falla
    -> ("", "", "") y el script sigue: nunca se rompe por un aviso puntual
    que no carga."""
    titulo = fila.get("titulo", "") or ""
    link = fila.get("link", "") or ""
    print(f"    [Detalle] leyendo: {titulo}")

    if not link:
        print("    [Detalle] error: sin link")
        return "", "", ""

    try:
        page = _get_page()
        page.goto(link, timeout=20000, wait_until="domcontentloaded")
        try:
            page.wait_for_selector("h1", timeout=6000)
        except Exception:
            pass
        texto = (page.inner_text("body") or "").strip()
    except Exception as e:
        print(f"    [Detalle] error: {e}")
        return "", "", ""

    if not texto:
        print("    [Detalle] error: página vacía")
        return "", "", ""

    # page.inner_text("body") trae tambien el menu/nav de arriba de la
    # pagina. El contenido real empieza donde aparece el titulo del aviso,
    # asi que se recorta ahi para que el resumen no arranque con ruido.
    if titulo:
        idx = texto.lower().find(titulo.lower())
        if idx > 0:
            texto = texto[idx:]

    resumen = _resumir_texto(texto)
    alertas = _detectar_alertas(texto)
    print("    [Detalle] ok")
    return texto, resumen, alertas


def ajustar_decision_por_descripcion(fila):
    """Usa descripcion_texto (ya cargado en 'fila') para ajustar
    decision_sugerida/motivo_decision/prioridad. Devuelve las 3, ajustadas
    o iguales si no hay nada relevante en el texto."""
    decision = fila.get("decision_sugerida", "REVISAR")
    motivo = fila.get("motivo_decision", "") or ""
    prioridad = fila.get("prioridad", 0) or 0
    texto = (fila.get("descripcion_texto") or "").lower()

    if not texto:
        return decision, motivo, prioridad

    ajustes = []
    hits_seniority = sum(1 for kw in _ALERTAS_SENIORIDAD if kw in texto)
    hits_skills = sum(1 for kw in _SKILLS_POSITIVOS if kw in texto)
    hits_junior = sum(1 for kw in _SIGNALS_JUNIOR if kw in texto)

    if hits_seniority >= 2 and decision != "DESCARTAR":
        ajustes.append(f"bajado a DESCARTAR: {hits_seniority} señales de seniority alta en la descripción")
        decision = "DESCARTAR"
    elif hits_seniority == 1 and decision == "POSTULAR":
        ajustes.append("bajado a REVISAR: la descripción pide algo de seniority que el título no dejaba ver")
        decision = "REVISAR"

    if hits_skills:
        prioridad += hits_skills
        ajustes.append(f"+{hits_skills} prioridad: matchea {hits_skills} skill(s) del perfil en la descripción")

    if hits_junior:
        prioridad += 1
        ajustes.append("+1 prioridad: la descripción pide perfil junior/trainee/estudiante")

    if hits_seniority >= 3 and hits_skills == 0:
        prioridad -= 2
        ajustes.append("-2 prioridad: muchos requisitos de seniority/excluyentes y ningún skill del perfil")

    prioridad = max(prioridad, 0)

    if ajustes:
        motivo = f"{motivo} | {'; '.join(ajustes)}" if motivo else "; ".join(ajustes)

    return decision, motivo, prioridad


# --- BANDEJA DE ACCIONES ---------------------------------------------------
# Solo SUGIERE que hacer con cada aviso. No postula, no hace clic en nada,
# no envia mensajes: arma la hoja ACCIONES para que la decision final la
# tome Sergio a mano.
ACCIONES_COLUMNAS = [
    "accion_sugerida", "estado_accion", "titulo", "empresa", "fuente",
    "categoria_detectada", "decision_sugerida", "prioridad", "motivo_decision",
    "alertas_aviso", "mensaje_postulacion", "link", "requiere_revision_manual",
]

# Alertas que ameritan frenar del todo (posible estafa, o requisito que
# claramente no se cumple). Si aparece alguna, la accion es NO ACCIONAR
# sin importar si la decision de base era POSTULAR o REVISAR.
_ALERTAS_GRAVES = [
    "pagar para postular", "curso pago", "capacitación paga", "capacitacion paga",
    "inscripción con costo", "inscripcion con costo", "zonajobs",
    "5 años", "6 años", "7 años", "senior",
]

# --- Alertas de la bandeja de ACCIONES (determinar_accion_sugerida) -----
# Listas propias de ESTA capa, separadas a propósito de _ALERTAS_GRAVES de
# arriba (esa la usa tambien _postulacion_es_segura, el gate del envio
# real de postulacion.py -- no se toca).
#
# Graves de verdad (estafa / portal externo / redirect): siempre
# NO ACCIONAR, sin importar decision_sugerida.
_ALERTAS_GRAVES_ACCION = [
    "pagar para postular", "curso pago", "capacitación paga", "capacitacion paga",
    "inscripción con costo", "inscripcion con costo", "zonajobs",
    "portal externo", "google forms", "hiringroom", "linkedin externo",
]

# Seniority/moderadas: en POSTULAR bajan a "REVISAR ANTES DE POSTULAR". En
# REVISAR no hacen nada extra -- REVISAR ya es la rama cautelosa, cae en
# "REVISAR MANUALMENTE" igual (antes esto se confundia con las graves de
# arriba y terminaba en NO ACCIONAR por error).
_ALERTAS_SENIORITY_ACCION = [
    "senior", "ssr", "semi senior", "semi-senior",
    "más de 3 años", "mas de 3 años", "4 años", "5 años", "6 años", "7 años",
    "contractor", "monotributo",
    "inglés avanzado", "ingles avanzado", "inglés c1", "ingles c1",
    "excluyente",
]

_MENSAJE_GENERICO_DEFAULT = (
    "Hola, me interesa postularme a la posición. Soy estudiante de Ingeniería en "
    "Informática y cuento con experiencia en áreas operativas, análisis de "
    "información, procesos y coordinación entre sectores. Me gustaría poder "
    "aportar desde un perfil responsable, analítico y con muchas ganas de seguir "
    "creciendo."
)

_MENSAJES_GENERICOS_POR_CATEGORIA = {
    "Data / BI": (
        "Hola, me interesa postularme a la posición. Soy estudiante de Ingeniería "
        "en Informática y cuento con experiencia en análisis de información, "
        "procesos y uso avanzado de Excel. Además, estoy fortaleciendo mis "
        "conocimientos en SQL, Python y Power BI. Me gustaría poder aportar desde "
        "un perfil analítico, ordenado y orientado a la mejora continua."
    ),
    "Soporte IT": (
        "Hola, me interesa postularme a la posición. Soy estudiante de Ingeniería "
        "en Informática y me interesa desarrollarme en soporte IT, resolución de "
        "problemas y atención a usuarios. Cuento con experiencia en coordinación "
        "operativa, análisis de información y seguimiento de procesos, además de "
        "conocimientos técnicos en sistemas y herramientas digitales."
    ),
    "Supply Chain": (
        "Hola, me interesa postularme a la posición. Cuento con experiencia en "
        "Supply Chain, planificación, compras, abastecimiento e inventarios, "
        "trabajando con Excel, análisis de información y coordinación entre "
        "sectores. Me gustaría aportar desde un perfil ordenado, analítico y "
        "orientado a procesos."
    ),
}


def determinar_accion_sugerida(fila):
    """Devuelve (accion_sugerida, estado_accion, mensaje_postulacion,
    requiere_revision_manual) para un aviso POSTULAR o REVISAR. Reglas
    simples, sin IA: no decide nada por si sola, solo ordena la bandeja."""
    decision = fila.get("decision_sugerida", "")
    alertas = str(fila.get("alertas_aviso", "") or "").lower()
    categoria = fila.get("categoria_detectada", "") or ""
    mensaje_existente = str(fila.get("mensaje_sugerido", "") or "").strip()

    hay_alerta_grave = any(kw in alertas for kw in _ALERTAS_GRAVES_ACCION)
    hay_alerta_seniority = any(kw in alertas for kw in _ALERTAS_SENIORITY_ACCION)

    if hay_alerta_grave:
        accion = "NO ACCIONAR"
    elif decision == "POSTULAR":
        accion = "REVISAR ANTES DE POSTULAR" if hay_alerta_seniority else "POSTULAR HOY"
    else:  # REVISAR (con o sin alertas de seniority, sigue siendo REVISAR MANUALMENTE)
        accion = "REVISAR MANUALMENTE"

    estado_accion = "pendiente"

    if mensaje_existente:
        mensaje_postulacion = mensaje_existente
    else:
        mensaje_postulacion = _MENSAJES_GENERICOS_POR_CATEGORIA.get(categoria, _MENSAJE_GENERICO_DEFAULT)

    requiere_revision_manual = "no" if accion == "NO ACCIONAR" else "sí"

    return accion, estado_accion, mensaje_postulacion, requiere_revision_manual


def construir_acciones(nuevas):
    """Arma el DataFrame de la hoja ACCIONES (solo POSTULAR/REVISAR; los
    DESCARTAR no entran) y el conteo por tipo de accion. No modifica
    'nuevas'. Nunca hace nada mas alla de armar esta tabla: no postula,
    no envia nada, no abre formularios."""
    conteo_base = {"POSTULAR HOY": 0, "REVISAR ANTES DE POSTULAR": 0,
                   "REVISAR MANUALMENTE": 0, "NO ACCIONAR": 0}

    acciones = nuevas[nuevas["decision_sugerida"].isin(["POSTULAR", "REVISAR"])].copy()
    if acciones.empty:
        return pd.DataFrame(columns=ACCIONES_COLUMNAS), conteo_base

    resultado = acciones.apply(lambda f: pd.Series(determinar_accion_sugerida(f)), axis=1)
    resultado.columns = ["accion_sugerida", "estado_accion", "mensaje_postulacion",
                          "requiere_revision_manual"]
    acciones = pd.concat([acciones.reset_index(drop=True), resultado.reset_index(drop=True)], axis=1)
    acciones = acciones[ACCIONES_COLUMNAS]

    conteo = conteo_base.copy()
    conteo.update(acciones["accion_sugerida"].value_counts().to_dict())
    return acciones, conteo


# --- POSTULACION ASISTIDA/AUTOMATICA (segura) -----------------------------
# Dominios propios donde SI tiene sentido intentar seguir el flujo de
# postulacion. Cualquier otro dominio (aunque el 'link' venga marcado como
# fuente Bumeran/Computrabajo por error) se trata como externo.
_DOMINIOS_POSTULACION_PROPIOS = ["bumeran.com.ar", "computrabajo.com"]

# Keywords de bloqueo (CAPTCHA/Cloudflare) reusadas de Bumeran/Indeed, para
# no duplicar la lista de deteccion.
_POSTULACION_BLOQUEO_KW = tuple(set(_BUMERAN_BLOQUEO_KW) | set(kw.lower() for kw in BLOQUEOS_INDEED))


def _postulacion_es_segura(fila):
    """Chequeo RAPIDO, sin abrir navegador: condiciones 1-4 del pedido
    (accion_sugerida, decision_sugerida, alertas graves, dominio propio).
    Devuelve (es_segura, motivo_si_no)."""
    if fila.get("accion_sugerida") != "POSTULAR HOY":
        return False, "accion_sugerida distinta de POSTULAR HOY"
    if fila.get("decision_sugerida") != "POSTULAR":
        return False, "decision_sugerida distinta de POSTULAR"

    alertas = str(fila.get("alertas_aviso", "") or "").lower()
    if any(kw in alertas for kw in _ALERTAS_GRAVES):
        return False, "hay alertas graves en el aviso"

    link = str(fila.get("link", "") or "").lower()
    if not any(dominio in link for dominio in _DOMINIOS_POSTULACION_PROPIOS):
        return False, "el link no pertenece a Bumeran ni Computrabajo"

    return True, ""


# --- ENVIO REAL (unico lugar del codigo que hace click en un boton de ---
# --- postulacion de verdad) ----------------------------------------------
# Textos candidatos para el boton de postulacion. Comparacion por texto
# EXACTO (recortado, sin mayusculas), no por substring: asi "Postularme"
# no matchea por error con "Postularme a otros empleos similares" o
# links de navegacion parecidos.
# "postulación rápida" / "postulacion rapida" se agregaron porque es el
# texto REAL verificado en vivo del boton de Bumeran (no estaba en la
# lista original: sin esto, la deteccion nunca encontraba nada en
# Bumeran). "Postularme" se confirmo en vivo como el texto real de
# Computrabajo, esa ya estaba en la lista.
_TEXTOS_BOTON_POSTULACION = [
    "postularme", "postular", "enviar postulación", "enviar postulacion",
    "enviar candidatura", "aplicar", "ya estoy interesado",
    "postulación rápida", "postulacion rapida",
]

# Señales de que hay algo que el agente no puede resolver solo (pregunta
# obligatoria, adjunto, CAPTCHA): si aparece cualquiera, no se envia nada.
_TEXTOS_BLOQUEO_FORMULARIO = [
    "pregunta", "responder", "adjuntar", "subir cv", "captcha",
    "verificación", "verificacion",
]

_TEXTOS_EXITO_POSTULACION = [
    "postulación enviada", "postulacion enviada", "te postulaste",
    "candidatura enviada", "recibimos tu postulación", "recibimos tu postulacion",
    "ya estás postulado", "ya estas postulado",
]


def detectar_boton_postulacion(page, fuente):
    """Busca el boton/enlace de postulacion en la pagina YA CARGADA, de
    forma conservadora. NUNCA hace click. Compara texto exacto (recortado,
    sin mayusculas) contra _TEXTOS_BOTON_POSTULACION sobre <button>, <a> y
    [role='button'] visibles. Devuelve (handle, motivo):
      (ElementHandle, "")  -> un solo candidato, claramente unico.
      (None, "motivo")     -> nada (0 candidatos) o ambiguo (2+
                               candidatos): nunca elige adivinando."""
    try:
        elementos = page.query_selector_all("button, a, [role='button']")
    except Exception as e:
        return None, f"No se pudieron leer los botones de la página: {e}"

    candidatos = []
    for el in elementos:
        try:
            if not el.is_visible():
                continue
            texto = (el.inner_text() or "").strip().lower()
        except Exception:
            continue
        if texto in _TEXTOS_BOTON_POSTULACION:
            candidatos.append((el, texto))

    if not candidatos:
        return None, (f"No se encontró ningún botón visible con alguno de estos "
                       f"textos: {', '.join(_TEXTOS_BOTON_POSTULACION)}. (fuente: {fuente})")
    if len(candidatos) > 1:
        textos = ", ".join(sorted(set(t for _, t in candidatos)))
        return None, (f"Se encontraron {len(candidatos)} botones candidatos "
                       f"({textos}): ambiguo, no se puede elegir sin arriesgar un "
                       f"click equivocado. (fuente: {fuente})")

    return candidatos[0][0], ""


def detectar_preguntas_o_bloqueos(page):
    """Revisa la pagina ACTUAL en busca de señales de que hay que
    completar algo que el agente no puede resolver solo: preguntas
    obligatorias, campos obligatorios sin completar, adjuntos, CAPTCHA.
    Nunca inventa una respuesta ni completa nada: si detecta cualquier
    señal, devuelve (True, motivo) para frenar el envio."""
    try:
        cuerpo = (page.inner_text("body") or "").lower()
    except Exception as e:
        return True, f"No se pudo leer la página para revisar preguntas/bloqueos: {e}"

    # "Preguntas frecuentes" es un link de footer presente en CASI TODAS
    # las paginas de Bumeran/Computrabajo (FAQ del sitio, no del aviso):
    # sin excluirlo, la palabra "pregunta" da falso positivo siempre y
    # bloquea el envio en cualquier pagina normal. Verificado en vivo.
    cuerpo_filtrado = cuerpo.replace("preguntas frecuentes", "")

    for kw in _TEXTOS_BLOQUEO_FORMULARIO:
        if kw in cuerpo_filtrado:
            return True, f"Se detectó texto de posible pregunta/bloqueo: '{kw}'."

    # campos obligatorios del formulario sin completar
    try:
        campos = page.query_selector_all(
            "form textarea[required], form input[required], form select[required]"
        )
        for campo in campos:
            try:
                valor = (campo.input_value() or "").strip()
            except Exception:
                valor = ""
            if not valor:
                return True, "Hay un campo obligatorio del formulario sin completar."
    except Exception:
        pass

    # preguntas/campos marcados con asterisco dentro de un formulario
    try:
        etiquetas = page.query_selector_all("form label, form span, form p, form legend")
        for etq in etiquetas:
            try:
                texto = (etq.inner_text() or "").strip()
            except Exception:
                continue
            if texto and (texto.endswith("*") or "(*)" in texto):
                return True, f"Se detectó una pregunta/campo marcado con asterisco: '{texto[:60]}'."
    except Exception:
        pass

    return False, ""


def _verificar_exito_postulacion(page):
    """Busca textos explicitos de exito tras un click de envio. Devuelve
    True SOLO si encuentra alguno de _TEXTOS_EXITO_POSTULACION; si no
    puede confirmarlo, devuelve False (nunca asume exito por default)."""
    try:
        cuerpo = (page.inner_text("body") or "").lower()
    except Exception:
        return False
    return any(kw in cuerpo for kw in _TEXTOS_EXITO_POSTULACION)


def enviar_postulacion_confirmada(page, fila):
    """UNICO lugar del codigo que hace click en un boton real de
    postulacion. Solo se llama desde _postular_oferta_logica() cuando ya
    se cumplieron TODAS las condiciones: MODO_EJECUCION='MANUAL',
    MODO_POSTULACION='CONFIRMADA', DRY_RUN_POSTULACION=False, la oferta
    paso los chequeos de seguridad (dominio propio, sin CAPTCHA, sin
    redirect externo), y el usuario ya confirmo una vez por consola.

    Aun asi, pide UNA confirmacion mas ("Última confirmación") justo
    antes de CUALQUIER click que pueda ser el envio definitivo -- tanto
    si es el primer boton (flujos de un solo paso, ej. "postulacion
    rapida") como si aparece un segundo boton de envio final (flujos de
    dos pasos). Nunca completa preguntas, nunca adjunta archivos, nunca
    saltea CAPTCHA: si detectar_preguntas_o_bloqueos() encuentra algo de
    eso en cualquier momento, se detiene y marca 'requiere revisión
    manual' sin clickear nada mas.

    Devuelve (estado_postulacion, resultado_postulacion,
    detalle_postulacion)."""
    titulo = fila.get("titulo", "")
    fuente = fila.get("fuente", "")

    boton, motivo = detectar_boton_postulacion(page, fuente)
    if boton is None:
        return ("requiere revisión manual", "",
                f"No se pudo identificar un botón único de postulación: {motivo}")

    hay_bloqueo, motivo_bloqueo = detectar_preguntas_o_bloqueos(page)
    if hay_bloqueo:
        return "requiere revisión manual", "", motivo_bloqueo

    print(f"\n  Se encontró el botón de postulación para: {titulo}")
    respuesta = input("Última confirmación: ¿enviar definitivamente esta postulación? [s/N] ")
    if respuesta not in ("s", "S"):
        return "omitido", "omitido por usuario", "El usuario no confirmó el envío final"

    try:
        boton.click()
        page.wait_for_timeout(2000)
        try:
            page.wait_for_load_state("networkidle", timeout=8000)
        except Exception:
            pass
    except Exception as e:
        return "error", "", f"Falló el click en el botón de postulación: {e}"

    hay_bloqueo2, motivo_bloqueo2 = detectar_preguntas_o_bloqueos(page)
    if hay_bloqueo2:
        return ("requiere revisión manual", "",
                f"Después del click apareció algo que no se puede completar solo: {motivo_bloqueo2}")

    # Flujo de dos pasos: si aparece un segundo boton de envio, pedir
    # confirmacion de nuevo antes de ese click (nunca se asume que el
    # primer click ya fue el envio definitivo).
    boton_final, _motivo_final = detectar_boton_postulacion(page, fuente)
    if boton_final is not None:
        print(f"\n  Apareció un segundo paso de postulación para: {titulo}")
        respuesta2 = input("Última confirmación: ¿enviar definitivamente esta postulación? [s/N] ")
        if respuesta2 not in ("s", "S"):
            return "omitido", "omitido por usuario", "El usuario no confirmó el envío final (segundo paso)"
        try:
            boton_final.click()
            page.wait_for_timeout(2000)
            try:
                page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass
        except Exception as e:
            return "error", "", f"Falló el click en el botón final de postulación: {e}"

    if _verificar_exito_postulacion(page):
        return "postulado", "éxito verificado", "Se encontró un mensaje de éxito tras el envío."

    return ("requiere revisión manual", "estado incierto",
            "Se hizo click pero no se pudo verificar mensaje de éxito.")


def _pedir_confirmacion_postulacion(fila):
    """Muestra el detalle del aviso por consola y pregunta si se confirma
    el envio. Solo devuelve True si la respuesta es exactamente 's' o 'S'
    (cualquier otra cosa, incluido enter en blanco, cuenta como no)."""
    print("\n" + "-" * 60)
    print(f"Título: {fila.get('titulo', '')}")
    print(f"Empresa: {fila.get('empresa', '')}")
    print(f"Fuente: {fila.get('fuente', '')}")
    print(f"Link: {fila.get('link', '')}")
    print(f"Categoría: {fila.get('categoria_detectada', '')}")
    print(f"Motivo decisión: {fila.get('motivo_decision', '')}")
    print(f"Alertas: {fila.get('alertas_aviso', '') or '(ninguna)'}")
    print(f"Mensaje sugerido: {fila.get('mensaje_postulacion', '')}")
    print("-" * 60)
    respuesta = input("¿Confirmás enviar esta postulación? [s/N] ")
    return respuesta in ("s", "S")


def _confirmada_permitida():
    """MODO_POSTULACION='CONFIRMADA' pregunta por consola (input()), asi
    que solo puede correr si hay alguien presente para contestar: en
    MANUAL siempre, en TEST solo si PROBAR_POSTULACION_EN_TEST=True (asi
    se puede probar el flujo sin ofertas nuevas reales). Nunca en AUTO ni
    en DEMO, sin excepcion."""
    if MODO_EJECUCION == "MANUAL":
        return True
    if MODO_EJECUCION == "TEST" and PROBAR_POSTULACION_EN_TEST:
        return True
    return False


def _postular_oferta_logica(fila):
    """Evalua UNA oferta ya filtrada como candidata (accion_sugerida =
    'POSTULAR HOY') y decide que hacer. Nunca inventa datos, nunca
    completa preguntas desconocidas, nunca saltea CAPTCHA, nunca postula
    en portales externos. Devuelve (estado_postulacion,
    resultado_postulacion, detalle_postulacion).

    Estados posibles: "postulado", "simulado", "requiere revisión manual",
    "omitido", "error".

    Nota importante sobre el envio REAL (click en el boton final de
    "Postularme"/"Enviar postulación"): en esta primera implementacion NO
    esta conectado en NINGUN modo, ni siquiera en CONFIRMADA despues de
    que el usuario dice que si. Postular es una accion irreversible hacia
    un tercero (el empleador), y el boton final varia de HTML en HTML
    (puede ser un <button type="submit">, un boton con texto "Enviar
    postulación"/"Postularme", o aparecer recien despues de un paso
    intermedio con preguntas) -- no hay un selector unico y confiable
    todavia para detectarlo en Bumeran y Computrabajo a la vez. Hasta que
    eso se identifique con seguridad, todas las ramas devuelven 'requiere
    revisión manual' en el paso final en vez de arriesgar un click mal
    dirigido. El modo 'ASISTIDA' (o 'CONFIRMADA' diciendo que si) deja el
    aviso verificado como seguro y listo para que el humano termine el
    envio a mano."""
    if MODO_POSTULACION == "OFF":
        return "omitido", "", "MODO_POSTULACION = OFF, no se procesó."

    if MODO_POSTULACION == "CONFIRMADA" and not _confirmada_permitida():
        # Nunca preguntar por consola (input()) sin alguien presente: en
        # AUTO/DEMO (y en TEST sin PROBAR_POSTULACION_EN_TEST) no hay
        # quien conteste y el script quedaria colgado esperando.
        return ("requiere revisión manual", "",
                "MODO_POSTULACION='CONFIRMADA' solo funciona con MODO_EJECUCION="
                "'MANUAL' (o 'TEST' con PROBAR_POSTULACION_EN_TEST=True). No se "
                f"pregunta ni se envía nada desde MODO_EJECUCION='{MODO_EJECUCION}'.")

    es_segura, motivo = _postulacion_es_segura(fila)
    if not es_segura:
        return "requiere revisión manual", "", f"No cumple condiciones de seguridad: {motivo}."

    link = fila.get("link", "")
    titulo = fila.get("titulo", "")

    try:
        page = _get_page()
        page.goto(link, timeout=20000, wait_until="domcontentloaded")
        try:
            page.wait_for_selector("h1", timeout=8000)
        except Exception:
            pass
        cuerpo = (page.inner_text("body") or "")
        cuerpo_low = cuerpo.lower()
    except Exception as e:
        return "error", "", f"No se pudo cargar el aviso: {e}"

    bloqueo = [kw for kw in _POSTULACION_BLOQUEO_KW if kw in cuerpo_low]
    if bloqueo:
        return "requiere revisión manual", "", f"CAPTCHA/Cloudflare detectado ({', '.join(bloqueo)})."

    externo = [kw for kw in BUMERAN_EXTERNOS_BLOQUEADOS if kw in cuerpo_low]
    if externo:
        return "requiere revisión manual", "", f"La postulación redirige a un portal externo ({', '.join(externo)})."

    # A partir de aca habria que reconocer el boton de postulacion y (en
    # ASISTIDA) llegar hasta el formulario, o (AUTO_SEGURO) revisar que no
    # pida preguntas obligatorias/adjuntos desconocidos antes de decidir.
    # Como esa parte varia mucho aviso a aviso y HTML a HTML, y como
    # completar un formulario ajeno con datos reales es justo el tipo de
    # accion que no se debe automatizar a ciegas, el agente se detiene
    # aca: llegó hasta el aviso, verificó que es seguro (propio, sin
    # bloqueo, sin redirect externo), y deja el resto para el humano.
    if MODO_POSTULACION == "ASISTIDA":
        return ("simulado", "",
                f"Oferta abierta y verificada como segura ({titulo}). "
                "El agente no completó ni envió el formulario: revisar y postular manualmente.")

    if MODO_POSTULACION == "CONFIRMADA":
        # Preguntas obligatorias desconocidas, adjuntos faltantes o datos
        # ambiguos: detectar_preguntas_o_bloqueos() (adentro de
        # enviar_postulacion_confirmada) trata cualquiera de esos casos
        # como "requiere revision". No se inventa nada.
        confirmado = _pedir_confirmacion_postulacion(fila)
        if not confirmado:
            return "omitido", "rechazado por usuario", "El usuario no confirmó el envío"

        if DRY_RUN_POSTULACION:
            return ("simulado", "",
                    "DRY RUN: confirmaste el envío, pero DRY_RUN_POSTULACION=True así "
                    "que no se hizo ningún click real.")

        if MODO_EJECUCION != "MANUAL":
            # Salvaguarda extra (ademas de _confirmada_permitida()): el
            # envio REAL con click de verdad solo puede pasar en una
            # corrida manual, nunca en TEST -- ni siquiera con
            # PROBAR_POSTULACION_EN_TEST=True, que solo habilita PROBAR
            # el flujo (con DRY_RUN=True), no enviar de verdad.
            return ("requiere revisión manual", "",
                    "DRY_RUN_POSTULACION=False solo puede enviar de verdad con "
                    f"MODO_EJECUCION='MANUAL'. Con MODO_EJECUCION='{MODO_EJECUCION}' "
                    "no se hace ningún click real aunque hayas confirmado.")

        # Unico camino de todo el archivo que puede terminar en un click
        # real: MANUAL + CONFIRMADA + DRY_RUN_POSTULACION=False + el
        # usuario ya confirmo una vez arriba. enviar_postulacion_confirmada
        # pide una confirmacion mas antes de cualquier click.
        return enviar_postulacion_confirmada(page, fila)

    # MODO_POSTULACION == "AUTO_SEGURO"
    if DRY_RUN_POSTULACION:
        return ("simulado", "",
                f"DRY RUN: la oferta pasó todos los chequeos de seguridad y se habría "
                "postulado. No se tocó ningún botón real.")

    return ("requiere revisión manual", "",
            "Envío automático real no está implementado (a propósito): postular es "
            "una acción irreversible hacia un tercero y requiere confirmación humana "
            "en el momento. Usar MODO_POSTULACION='ASISTIDA' para hacerlo con ayuda "
            "del agente, o postular manualmente.")


def postular_oferta(fila, es_test=False):
    """Punto de entrada publico: corre _postular_oferta_logica() y, si
    es_test=True (viene de MODO_EJECUCION='TEST' con
    PROBAR_POSTULACION_EN_TEST=True), marca claramente el resultado como
    una prueba -- nunca se confunde con una postulacion real en el log ni
    en la hoja ACCIONES. No cambia ninguna otra logica: los mismos
    chequeos de seguridad, el mismo tope, la misma confirmacion si
    corresponde."""
    estado, resultado_txt, detalle = _postular_oferta_logica(fila)
    if es_test:
        etiqueta = "[PRUEBA MODO TEST]"
        resultado_txt = f"{etiqueta} {resultado_txt}".strip() if resultado_txt else etiqueta
    return estado, resultado_txt, detalle


def ejecutar_postulaciones(df_acciones, es_test=False):
    """Recorre la hoja ACCIONES, filtra 'POSTULAR HOY', respeta
    MAX_POSTULACIONES_POR_CORRIDA, y llama a postular_oferta() por cada
    una. Registra cada intento en postulaciones_log.xlsx. Devuelve una
    COPIA de df_acciones con 3 columnas nuevas (estado_postulacion,
    resultado_postulacion, detalle_postulacion). Si MODO_POSTULACION es
    'OFF', no toca el navegador ni procesa nada: vuelve con esas 3
    columnas vacías.

    es_test=True (llamado desde MODO_EJECUCION='TEST' con
    PROBAR_POSTULACION_EN_TEST=True) NO cambia ningun chequeo de
    seguridad: solo hace que cada resultado quede marcado como prueba
    (via postular_oferta(..., es_test=True)) y no toca vistos.json ni
    historial_trabajos.xlsx -- eso ya esta separado de esta funcion, en
    main(), y sigue igual sin importar este flag."""
    resultado = df_acciones.copy()
    resultado["estado_postulacion"] = ""
    resultado["resultado_postulacion"] = ""
    resultado["detalle_postulacion"] = ""

    if resultado.empty or MODO_POSTULACION == "OFF":
        return resultado

    if MODO_POSTULACION == "CONFIRMADA" and not _confirmada_permitida():
        # Defensa en profundidad: ni siquiera abrir el navegador. La misma
        # regla se re-chequea adentro de postular_oferta() por si esta
        # funcion se llama de otro lado en el futuro.
        print(f"\n  MODO_POSTULACION='CONFIRMADA' requiere MODO_EJECUCION='MANUAL' "
              f"(o 'TEST' con PROBAR_POSTULACION_EN_TEST=True). Como MODO_EJECUCION="
              f"'{MODO_EJECUCION}', no se pregunta ni se envía nada esta corrida.")
        return resultado

    candidatas_idx = resultado.index[resultado["accion_sugerida"] == "POSTULAR HOY"]
    candidatas_idx = candidatas_idx[:MAX_POSTULACIONES_POR_CORRIDA]

    if len(candidatas_idx) == 0:
        return resultado

    print(f"\n--- POSTULACIÓN ({MODO_POSTULACION}"
          f"{', DRY RUN' if DRY_RUN_POSTULACION else ''}"
          f"{', MODO TEST' if es_test else ''}) ---")
    print(f"  Procesando {len(candidatas_idx)} oferta(s) POSTULAR HOY "
          f"(tope: {MAX_POSTULACIONES_POR_CORRIDA})")
    if es_test:
        print("  MODO TEST: no se toca vistos.json ni historial_trabajos.xlsx, "
              "y no se asume éxito en ningún caso.")

    for idx in candidatas_idx:
        fila = resultado.loc[idx]
        titulo = fila.get("titulo", "")
        print(f"  [Postulación] procesando: {titulo}")

        estado, resultado_txt, detalle = postular_oferta(fila, es_test=es_test)
        print(f"  [Postulación] {estado}: {detalle}")

        resultado.at[idx, "estado_postulacion"] = estado
        resultado.at[idx, "resultado_postulacion"] = resultado_txt
        resultado.at[idx, "detalle_postulacion"] = detalle

        _registrar_postulacion_log({
            "fecha_hora": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "titulo": titulo,
            "empresa": fila.get("empresa", ""),
            "fuente": fila.get("fuente", ""),
            "link": fila.get("link", ""),
            "accion_sugerida": fila.get("accion_sugerida", ""),
            "estado_postulacion": estado,
            "resultado_postulacion": resultado_txt,
            "detalle_postulacion": detalle,
        })

    return resultado


# --- FORMATO DEL EXCEL ----------------------------------------------------
# Colores suaves (no agresivos), estilo "semaforo" clasico de Excel.
_COLORES_DECISION = {
    "POSTULAR": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "REVISAR": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "DESCARTAR": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
}


def aplicar_formato_excel(writer, df, resumen):
    """Formato visual de las hojas de datos (TODOS, una por fuente, y
    ACCIONES): fila 1 congelada, autofiltro, ancho de columna aproximado,
    link como hipervinculo, y relleno suave de fila segun
    decision_sugerida. No toca la hoja RESUMEN (esa tiene su propio
    layout, armado aparte).

    Los indices de 'link'/'decision_sugerida' se calculan POR HOJA (leyendo
    su propia fila de encabezados), no del 'df' recibido: ACCIONES tiene un
    orden de columnas distinto al de TODOS/por-fuente, y esto evita que se
    le aplique el hipervinculo o el color a la columna equivocada."""
    wb = writer.book

    for nombre_hoja in wb.sheetnames:
        if nombre_hoja == "RESUMEN":
            continue
        ws = wb[nombre_hoja]
        if ws.max_row < 2:
            continue

        encabezados = [c.value for c in ws[1]]
        col_link_idx = encabezados.index("link") + 1 if "link" in encabezados else None
        col_decision_idx = (encabezados.index("decision_sugerida") + 1
                             if "decision_sugerida" in encabezados else None)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ancho de columna aproximado (openpyxl no tiene autofit real)
        for col_cells in ws.columns:
            largo = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            letra = col_cells[0].column_letter
            ws.column_dimensions[letra].width = min(max(largo + 2, 10), 60)

        if col_link_idx:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_link_idx)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.value = "Abrir oferta"
                    cell.style = "Hyperlink"

        # relleno de fila segun decision_sugerida (despues del hipervinculo,
        # para no perder el fill cuando se aplica el estilo "Hyperlink")
        if col_decision_idx:
            for row in range(2, ws.max_row + 1):
                valor_decision = ws.cell(row=row, column=col_decision_idx).value
                relleno = _COLORES_DECISION.get(str(valor_decision))
                if relleno:
                    for c in ws[row]:
                        c.fill = relleno


def crear_hoja_resumen(writer, datos_resumen):
    """Crea la hoja RESUMEN con lo esencial de la corrida + el top 5. Se
    reordena a la primera posicion despues, en main(). No recalcula nada:
    solo vuelca lo que ya viene calculado en datos_resumen."""
    wb = writer.book
    ws = wb.create_sheet("RESUMEN")

    titulo_font = Font(bold=True, size=14)
    seccion_font = Font(bold=True, size=11)
    etiqueta_font = Font(bold=True)

    es_demo = bool(datos_resumen.get("es_demo"))
    titulo_hoja = "Buscador de trabajos - Resumen de la corrida"
    if es_demo:
        titulo_hoja += " — MODO DEMO"
    ws["A1"] = titulo_hoja
    ws["A1"].font = titulo_font
    ws.merge_cells("A1:E1")

    fila = 3
    if es_demo:
        campos = [
            ("Fecha/hora de ejecución", datos_resumen["fecha_hora"]),
            ("Modo de ejecución", "MODO DEMO — sin scraping, sin tocar historial"),
            ("Duración total", datos_resumen["duracion"]),
            ("Archivo generado", datos_resumen["archivo"]),
            ("Ofertas demo generadas", datos_resumen["cant_nuevos"]),
            ("POSTULAR", datos_resumen["postular"]),
            ("REVISAR", datos_resumen["revisar"]),
            ("DESCARTAR", datos_resumen["descartar"]),
        ]
    else:
        campos = [
            ("Fecha/hora de ejecución", datos_resumen["fecha_hora"]),
            ("Modo de ejecución", datos_resumen["modo_ejecucion"]),
            ("Duración total", datos_resumen["duracion"]),
            ("Archivo generado", datos_resumen["archivo"]),
            ("Ofertas nuevas", datos_resumen["cant_nuevos"]),
            ("Ofertas ya vistas", datos_resumen["cant_repetidos"]),
            ("POSTULAR", datos_resumen["postular"]),
            ("REVISAR", datos_resumen["revisar"]),
            ("DESCARTAR", datos_resumen["descartar"]),
        ]
        if datos_resumen.get("es_test"):
            campos.insert(3, ("Modo", "MODO TEST — no se actualizó el historial"))

    for etiqueta, valor in campos:
        ws.cell(row=fila, column=1, value=etiqueta).font = etiqueta_font
        ws.cell(row=fila, column=2, value=valor)
        fila += 1

    fila += 1
    ws.cell(row=fila, column=1, value="Top recomendadas").font = seccion_font
    fila += 1

    for col, encabezado in enumerate(["#", "Título", "Empresa", "Fuente", "Decisión", "Categoría"], start=1):
        ws.cell(row=fila, column=col, value=encabezado).font = etiqueta_font
    fila += 1

    top5 = datos_resumen.get("top5") or []
    if not top5:
        ws.cell(row=fila, column=1, value="(sin ofertas)")
    for item in top5:
        ws.cell(row=fila, column=1, value=item["n"])
        ws.cell(row=fila, column=2, value=item["titulo"])
        ws.cell(row=fila, column=3, value=item["empresa"])
        ws.cell(row=fila, column=4, value=item["fuente"])
        ws.cell(row=fila, column=5, value=item["decision"])
        ws.cell(row=fila, column=6, value=item.get("categoria", ""))
        relleno = _COLORES_DECISION.get(item["decision"])
        if relleno:
            for col in range(1, 7):
                ws.cell(row=fila, column=col).fill = relleno
        fila += 1

    fila += 1
    ws.cell(row=fila, column=1, value="Acciones sugeridas").font = seccion_font
    fila += 1
    conteo_acciones = datos_resumen.get("conteo_acciones") or {}
    for etiqueta in ("POSTULAR HOY", "REVISAR ANTES DE POSTULAR",
                      "REVISAR MANUALMENTE", "NO ACCIONAR"):
        ws.cell(row=fila, column=1, value=etiqueta).font = etiqueta_font
        ws.cell(row=fila, column=2, value=conteo_acciones.get(etiqueta, 0))
        fila += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 22


# ----------------------------------------------------------------------
# MODO DEMO (sin internet, sin tocar historial)
# ----------------------------------------------------------------------
def generar_datos_demo():
    """6 ofertas simuladas, con el mismo formato que devuelven bumeran()/
    computrabajo() (relevancia, titulo, empresa, ubicacion, modalidad,
    fecha, link, busqueda, fuente) + una descripcion larga simulada
    (solo interna, no se guarda tal cual en el Excel) para que pasen por
    ajustar_decision_por_descripcion() igual que un aviso real. Pensadas
    para cubrir: 2 POSTULAR, 2 REVISAR, 2 DESCARTAR (una por seniority,
    otra por rubro no deseado)."""
    return [
        {
            "titulo": "Analista de Datos Junior", "empresa": "Demo Analytics",
            "ubicacion": "CABA", "modalidad": "Híbrido", "fecha": "Publicado hoy",
            "link": "https://demo.local/ofertas/analista-de-datos-junior",
            "busqueda": "analista de datos", "fuente": "Demo",
            "descripcion_texto_demo": (
                "Buscamos Analista de Datos Junior para sumarse al equipo de Business "
                "Intelligence. Requisitos: manejo de SQL, Excel avanzado, Power BI y "
                "conocimientos básicos de Python. Valoramos perfiles estudiantes "
                "avanzados o junior con ganas de aprender y crecer en el área de datos."
            ),
        },
        {
            "titulo": "Soporte IT Mesa de Ayuda Junior", "empresa": "Demo IT Services",
            "ubicacion": "Buenos Aires", "modalidad": "Presencial", "fecha": "Publicado ayer",
            "link": "https://demo.local/ofertas/soporte-it-mesa-de-ayuda-junior",
            "busqueda": "soporte it", "fuente": "Demo",
            "descripcion_texto_demo": (
                "Se busca Analista de Soporte IT / Mesa de Ayuda para gestión de "
                "tickets, resolución de incidentes de hardware y software, y atención "
                "a usuarios internos. Buen ambiente de trabajo, capacitación constante. "
                "Perfil junior, con ganas de aprender."
            ),
        },
        {
            "titulo": "Analista Funcional", "empresa": "Demo Software Factory",
            "ubicacion": "CABA", "modalidad": "Remoto", "fecha": "Publicado hace 2 días",
            "link": "https://demo.local/ofertas/analista-funcional",
            "busqueda": "analista funcional", "fuente": "Demo",
            "descripcion_texto_demo": (
                "Analista Funcional para relevamiento de requerimientos, documentación "
                "funcional, redacción de user stories y participación en instancias de "
                "UAT junto al equipo de negocio. Se valora experiencia previa en "
                "proyectos de software, aunque no es excluyente."
            ),
        },
        {
            "titulo": "Analista de Supply Chain", "empresa": "Demo Logística",
            "ubicacion": "GBA", "modalidad": "Presencial", "fecha": "Publicado hace 1 día",
            "link": "https://demo.local/ofertas/analista-de-supply-chain",
            "busqueda": "supply chain", "fuente": "Demo",
            "descripcion_texto_demo": (
                "Buscamos Analista de Supply Chain para el área de compras e "
                "inventario. Tareas: planificación de abastecimiento, seguimiento de "
                "pedidos, manejo de Excel y coordinación operativa con depósito y "
                "logística."
            ),
        },
        {
            "titulo": "Data Engineer Senior", "empresa": "Demo Tech Senior",
            "ubicacion": "CABA", "modalidad": "Híbrido", "fecha": "Publicado hoy",
            "link": "https://demo.local/ofertas/data-engineer-senior",
            "busqueda": "data engineer", "fuente": "Demo",
            "descripcion_texto_demo": (
                "Data Engineer Senior con más de 5 años de experiencia excluyente en "
                "arquitectura de datos cloud (AWS/GCP), pipelines de datos a gran "
                "escala e inglés avanzado excluyente. Seniority alta, liderazgo de "
                "equipo técnico."
            ),
        },
        {
            "titulo": "Vendedor Telefónico Call Center", "empresa": "Demo Ventas",
            "ubicacion": "CABA", "modalidad": "Presencial", "fecha": "Publicado hoy",
            "link": "https://demo.local/ofertas/vendedor-telefonico-call-center",
            "busqueda": "vendedor call center", "fuente": "Demo",
            "descripcion_texto_demo": (
                "Se buscan Vendedores Telefónicos para call center de venta de "
                "productos. Objetivos comerciales diarios, comisiones por ventas. No "
                "se requiere experiencia previa."
            ),
        },
    ]


def _ejecutar_modo_demo(inicio_ejecucion):
    """Genera el Excel de demostracion: sin scraping, sin tocar
    vistos.json/historial_trabajos.xlsx. Reusa las mismas funciones de
    clasificacion/ficha/categoria/formato que el resto del script."""
    print("MODO DEMO: generando ofertas simuladas (sin conexión a internet)...")

    filas_demo = generar_datos_demo()

    # pd.DataFrame(filas_demo) crea una columna por cada key de los dicts
    # (incluye "descripcion_texto_demo", que es solo para uso interno). Al
    # seleccionar nuevas[list(COLUMNAS)] mas abajo, esa columna se descarta
    # y queda el mismo formato/orden que un Excel real.
    nuevas = pd.DataFrame(filas_demo)
    nuevas["relevancia"] = [
        puntuar({"_texto": f"{f['titulo']} {f['empresa']}"}) for f in filas_demo
    ]
    nuevas = nuevas[list(COLUMNAS)]

    # decision base (igual que un aviso real) + ajuste por la descripcion
    # simulada (misma funcion que usa el analisis de detalle real)
    decisiones, motivos, prioridades = [], [], []
    resumenes, alertas_list = [], []
    for f in filas_demo:
        decision, motivo, prioridad = clasificar_decision(f)
        texto_demo = f["descripcion_texto_demo"]
        resumen = _resumir_texto(texto_demo)
        alertas = _detectar_alertas(texto_demo)

        fila_ajuste = {
            "decision_sugerida": decision, "motivo_decision": motivo,
            "prioridad": prioridad, "descripcion_texto": texto_demo,
        }
        decision, motivo, prioridad = ajustar_decision_por_descripcion(fila_ajuste)

        decisiones.append(decision)
        motivos.append(motivo)
        prioridades.append(prioridad)
        resumenes.append(resumen)
        alertas_list.append(alertas)

    nuevas["decision_sugerida"] = decisiones
    nuevas["motivo_decision"] = motivos
    nuevas["prioridad"] = prioridades
    nuevas["descripcion_resumen"] = resumenes
    nuevas["alertas_aviso"] = alertas_list

    nuevas["_orden_decision"] = nuevas["decision_sugerida"].map(_ORDEN_DECISION)
    nuevas = nuevas.sort_values(
        by=["_orden_decision", "prioridad", "relevancia"],
        ascending=[True, False, False],
    ).drop(columns=["_orden_decision"])

    conteo_decision = nuevas["decision_sugerida"].value_counts().to_dict()
    print("\n--- DIAGNOSTICO: decision sugerida (DEMO) ---")
    print(f"  POSTULAR: {conteo_decision.get('POSTULAR', 0)}")
    print(f"  REVISAR: {conteo_decision.get('REVISAR', 0)}")
    print(f"  DESCARTAR: {conteo_decision.get('DESCARTAR', 0)}")

    def _ficha_fila(fila):
        if fila["decision_sugerida"] == "POSTULAR":
            return pd.Series(generar_ficha_postulacion(fila))
        if fila["decision_sugerida"] == "REVISAR":
            return pd.Series(("", "", "", "Revisar manualmente"))
        return pd.Series(("", "", "", "No priorizar"))

    nuevas[["porque_conviene", "encaje_perfil", "mensaje_sugerido", "accion_recomendada"]] = (
        nuevas.apply(_ficha_fila, axis=1)
    )
    nuevas["categoria_detectada"] = nuevas.apply(detectar_categoria, axis=1)
    cant_postular = conteo_decision.get("POSTULAR", 0)

    top5_lineas, top5_items = [], []
    for i, (_, fila) in enumerate(nuevas.head(5).iterrows(), start=1):
        top5_lineas.append(
            f"{i}. {fila['titulo']} - {fila['empresa']} - "
            f"{fila['fuente']} - {fila['decision_sugerida']} - {fila['categoria_detectada']}"
        )
        top5_items.append({
            "n": i, "titulo": fila["titulo"], "empresa": fila["empresa"],
            "fuente": fila["fuente"], "decision": fila["decision_sugerida"],
            "categoria": fila["categoria_detectada"],
        })
    top5_texto = "\n".join(top5_lineas) if top5_lineas else "(sin ofertas)"

    # --- bandeja de acciones: solo sugiere, no postula ni envia nada -----
    # (los links de DEMO son ficticios: no se procesa postulacion sobre
    # datos simulados, solo se dejan las columnas vacias para mantener
    # el mismo esquema que un Excel real)
    acciones_df, conteo_acciones = construir_acciones(nuevas)
    acciones_df["estado_postulacion"] = ""
    acciones_df["resultado_postulacion"] = ""
    acciones_df["detalle_postulacion"] = ""

    archivo = construir_ruta_excel(es_test=False, es_demo=True)

    datos_resumen = {
        "fecha_hora": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "modo_ejecucion": "DEMO",
        "duracion": _formatear_duracion(inicio_ejecucion),
        "archivo": str(archivo),
        "cant_nuevos": len(nuevas),
        "cant_repetidos": 0,
        "postular": conteo_decision.get("POSTULAR", 0),
        "revisar": conteo_decision.get("REVISAR", 0),
        "descartar": conteo_decision.get("DESCARTAR", 0),
        "es_test": False,
        "es_demo": True,
        "top5": top5_items,
        "conteo_acciones": conteo_acciones,
    }

    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        crear_hoja_resumen(writer, datos_resumen)
        nuevas.to_excel(writer, sheet_name="TODOS", index=False)
        nuevas.to_excel(writer, sheet_name="Demo", index=False)
        acciones_df.to_excel(writer, sheet_name="ACCIONES", index=False)
        aplicar_formato_excel(writer, nuevas, datos_resumen)

        wb = writer.book
        if "RESUMEN" in wb.sheetnames:
            idx = wb.sheetnames.index("RESUMEN")
            if idx != 0:
                wb.move_sheet("RESUMEN", offset=-idx)
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"]["A1"].value is None:
            wb.remove(wb["Sheet"])

    print(f"\nListo (DEMO): {archivo}")
    print(f"  {len(nuevas)} ofertas demo generadas.")
    print("  No se ejecutó scraping. No se actualizó historial.")

    mensaje_final = (
        f"Buscador terminado en MODO DEMO.\n\n"
        f"Archivo generado:\n{archivo}\n\n"
        f"Ofertas demo generadas: {len(nuevas)}\n"
        f"No se ejecutó scraping.\n"
        f"No se actualizó historial.\n\n"
        f"POSTULAR: {conteo_decision.get('POSTULAR', 0)}\n"
        f"REVISAR: {conteo_decision.get('REVISAR', 0)}\n"
        f"DESCARTAR: {conteo_decision.get('DESCARTAR', 0)}\n\n"
        f"Top recomendadas:\n{top5_texto}"
    )
    if cant_postular > 0:
        mensaje_final += (
            f"\n\nSe generaron fichas de postulación para {cant_postular} "
            f"ofertas marcadas como POSTULAR."
        )
    mensaje_final += (
        f"\n\nAcciones sugeridas:\n"
        f"POSTULAR HOY: {conteo_acciones.get('POSTULAR HOY', 0)}\n"
        f"REVISAR ANTES DE POSTULAR: {conteo_acciones.get('REVISAR ANTES DE POSTULAR', 0)}\n"
        f"REVISAR MANUALMENTE: {conteo_acciones.get('REVISAR MANUALMENTE', 0)}\n"
        f"NO ACCIONAR: {conteo_acciones.get('NO ACCIONAR', 0)}"
    )
    mensaje_final += f"\n\nDuración total: {_formatear_duracion(inicio_ejecucion)}."

    abrir_excel_si_corresponde(archivo)
    avisar_fin(mensaje_final)


def main():
    inicio_ejecucion = datetime.datetime.now()

    global PERFIL_USUARIO_TEXTO
    PERFIL_USUARIO_TEXTO = cargar_perfil_usuario()

    print("\n--- MODO DE USO ---")
    print(f"Modo ejecución: {MODO_EJECUCION}")
    print(f"Navegador visible: {NAVEGADOR_VISIBLE}")
    print(f"Indeed activo: {ACTIVAR_INDEED}")
    print(f"Abrir Excel al final: {ABRIR_EXCEL_AL_FINAL}")
    print(f"Mostrar cartel final: {MOSTRAR_CARTEL_FINAL}")
    print(f"Modo test con repetidas: {MODO_TEST_GENERAR_EXCEL_AUNQUE_SEAN_REPETIDAS}")
    print(f"Descartar externos Bumeran: {DESCARTAR_BUMERAN_EXTERNOS}")
    print(f"Días máximos de publicación: {CT_DIAS}")
    print(f"Sonido final: {SONIDO_FINAL}")
    print(f"Perfil usuario cargado: {'sí' if PERFIL_USUARIO_TEXTO else 'no'}")
    print(f"Analizar descripción detalle: {ANALIZAR_DESCRIPCION_DETALLE}")
    print(f"Modo postulación: {MODO_POSTULACION}")
    print(f"Dry run postulación: {DRY_RUN_POSTULACION}")
    print(f"Máximo postulaciones por corrida: {MAX_POSTULACIONES_POR_CORRIDA}")
    print(f"Probar postulación en TEST: {PROBAR_POSTULACION_EN_TEST}")
    print("-------------------\n")

    if MODO_EJECUCION == "DEMO":
        # Nada de scraping ni de historial: se corta aca mismo, antes de
        # tocar Bumeran/Computrabajo/Indeed/vistos.json.
        _ejecutar_modo_demo(inicio_ejecucion)
        return

    # --- ETAPA 1: traer resultados CRUDOS, sin filtrar nada todavia ----
    todas_crudas = []
    crudos_por_fuente = {}
    for funcion in FUENTES:
        total_fuente = 0
        for q in BUSQUEDAS:
            print(f"Buscando '{q}' en {funcion.__name__}...")
            filas = funcion(q)
            total_fuente += len(filas)
            todas_crudas.extend(filas)
        crudos_por_fuente[funcion.__name__] = total_fuente

    _cerrar_navegador()  # cerrar Chrome de Playwright (Bumeran/Indeed)

    print("\n--- DIAGNOSTICO: resultados crudos por fuente (sin filtrar) ---")
    for nombre, cant in crudos_por_fuente.items():
        print(f"  {nombre}: {cant}")
    if crudos_por_fuente.get("indeed") == 0 and _INDEED_BLOQUEOS_DETECTADOS > 0:
        print("\nIndeed devolvió 0 porque fue bloqueado por Cloudflare.")
        print("No es un problema de selectores.")

    if not todas_crudas:
        print("\nNinguna fuente devolvio resultados crudos. El problema esta "
              "en el scraping (selectores/bloqueo), no en el filtro. Revisar "
              "los logs [Bumeran]/[Indeed] de arriba.")
        avisar_fin(f"El buscador terminó sin resultados. "
                   f"Duración: {_formatear_duracion(inicio_ejecucion)}.")
        return

    # --- ETAPA 2: filtro de zona ---------------------------------------
    todas, descartadas_zona = [], 0
    for fila in todas_crudas:
        if _pasa_zona(fila.get("ubicacion")):
            todas.append(fila)
        else:
            descartadas_zona += 1
    print(f"Tras filtro de zona: {len(todas)} (descartadas {descartadas_zona})")

    if not todas:
        print("\nSin resultados tras el filtro de zona.")
        print(f"(Se descartaron {descartadas_zona} por estar fuera de zona.)")
        avisar_fin(f"El buscador terminó sin resultados. "
                   f"Duración: {_formatear_duracion(inicio_ejecucion)}.")
        return

    for fila in todas:
        fila["relevancia"] = puntuar(fila)
        fila.pop("_texto", None)

    df = pd.DataFrame(todas, columns=COLUMNAS)

    # --- ETAPA 3: FILTRO FINO -------------------------------------------
    antes = len(df)
    # 1) relevancia minima
    if MIN_RELEVANCIA > 0:
        df = df[df["relevancia"] >= MIN_RELEVANCIA]
    print(f"Tras filtro de relevancia (>= {MIN_RELEVANCIA}): {len(df)} "
          f"(descartadas {antes - len(df)})")

    # 2) el titulo debe contener alguna palabra de la lista
    antes_titulo = len(df)
    if TITULO_DEBE_CONTENER:
        partes = []
        for p in TITULO_DEBE_CONTENER:
            pl = re.escape(p.lower())
            # palabras cortas/ambiguas: exigir limites de palabra
            if p.lower() in TITULO_PALABRA_ENTERA:
                partes.append(rf"\b{pl}\b")
            else:
                partes.append(pl)
        patron = "|".join(partes)
        df = df[df["titulo"].fillna("").str.lower().str.contains(patron, regex=True)]
    print(f"Tras filtro de titulo (debe contener): {len(df)} "
          f"(descartadas {antes_titulo - len(df)})")

    # 3) lista negra: descartar titulos con palabras de otros rubros
    antes_negra = len(df)
    if TITULO_NO_DEBE_CONTENER:
        patron_neg = "|".join(re.escape(p.lower()) for p in TITULO_NO_DEBE_CONTENER)
        df = df[~df["titulo"].fillna("").str.lower().str.contains(patron_neg, regex=True)]
    print(f"Tras lista negra: {len(df)} (descartadas {antes_negra - len(df)})")

    filtradas = antes - len(df)

    # --- DEDUP por titulo+empresa (el link trae sufijos que enganan
    #     a drop_duplicates, asi que deduplicamos por contenido) -------
    df["_clave"] = (df["titulo"].fillna("").str.lower().str.strip() + "|" +
                    df["empresa"].fillna("").str.lower().str.strip())
    dup = len(df)
    df = df.drop_duplicates(subset=["_clave"]).drop(columns=["_clave"])
    duplicadas = dup - len(df)

    df = df.sort_values("relevancia", ascending=False)

    if df.empty:
        print("\nTodo quedo filtrado. Baja MIN_RELEVANCIA o vacia "
              "TITULO_DEBE_CONTENER para ver mas.")
        avisar_fin(f"El buscador terminó sin resultados. "
                   f"Duración: {_formatear_duracion(inicio_ejecucion)}.")
        return

    # --- HISTORIAL: separar avisos nuevos de los ya vistos --------------
    # Clave = link. Si no existe vistos.json o esta vacio, no rompe: todo
    # cuenta como nuevo (se crea el historial desde cero).
    vistos = _cargar_vistos()
    links_vistos = set(vistos.keys())
    df["_ya_visto"] = df["link"].fillna("").isin(links_vistos)
    repetidas = df[df["_ya_visto"]].drop(columns=["_ya_visto"])
    nuevas = df[~df["_ya_visto"]].drop(columns=["_ya_visto"])

    cant_repetidos = len(repetidas)
    cant_nuevos = len(nuevas)

    print("\n--- DIAGNOSTICO: historial ---")
    print(f"  ya vistas antes (repetidas): {cant_repetidos}")
    print(f"  nuevas: {cant_nuevos}")

    es_test = False
    if nuevas.empty:
        if MODO_TEST_GENERAR_EXCEL_AUNQUE_SEAN_REPETIDAS and not repetidas.empty:
            es_test = True
            print(f"\nMODO TEST: 0 ofertas nuevas. Uso las {cant_repetidos} repetidas "
                  "para generar un Excel de prueba (no se toca el historial).")
            nuevas = repetidas.copy()
        else:
            print("\nNo hay ofertas nuevas: todo lo encontrado ya estaba en el historial.")
            avisar_fin(f"Buscador terminó: 0 ofertas nuevas ({cant_repetidos} ya vistas). "
                       f"Duración: {_formatear_duracion(inicio_ejecucion)}.")
            return

    # --- DECISION SUGERIDA: clasificar cada aviso nuevo, sin descartar --
    nuevas[["decision_sugerida", "motivo_decision", "prioridad"]] = nuevas.apply(
        lambda fila: pd.Series(clasificar_decision(fila)), axis=1
    )

    # --- ANALISIS DE DESCRIPCION: solo POSTULAR/REVISAR, y solo si esta
    #     prendido el interruptor (puede ser lento: entra a cada aviso). ---
    def _detalle_fila(fila):
        if not ANALIZAR_DESCRIPCION_DETALLE or fila["decision_sugerida"] not in ("POSTULAR", "REVISAR"):
            return pd.Series((fila["decision_sugerida"], fila["motivo_decision"],
                               fila["prioridad"], "", ""))

        texto, resumen, alertas = extraer_detalle_aviso(fila)
        if not texto:
            return pd.Series((fila["decision_sugerida"], fila["motivo_decision"],
                               fila["prioridad"], resumen, alertas))

        fila_ajuste = dict(fila)
        fila_ajuste["descripcion_texto"] = texto
        decision, motivo, prioridad = ajustar_decision_por_descripcion(fila_ajuste)
        return pd.Series((decision, motivo, prioridad, resumen, alertas))

    nuevas[["decision_sugerida", "motivo_decision", "prioridad",
            "descripcion_resumen", "alertas_aviso"]] = nuevas.apply(_detalle_fila, axis=1)

    # orden final (recalculado por si el analisis de descripcion cambio
    # alguna decision o prioridad)
    nuevas["_orden_decision"] = nuevas["decision_sugerida"].map(_ORDEN_DECISION)
    nuevas = nuevas.sort_values(
        by=["_orden_decision", "prioridad", "relevancia"],
        ascending=[True, False, False],
    ).drop(columns=["_orden_decision"])

    conteo_decision = nuevas["decision_sugerida"].value_counts().to_dict()
    print("\n--- DIAGNOSTICO: decision sugerida ---")
    print(f"  POSTULAR: {conteo_decision.get('POSTULAR', 0)}")
    print(f"  REVISAR: {conteo_decision.get('REVISAR', 0)}")
    print(f"  DESCARTAR: {conteo_decision.get('DESCARTAR', 0)}")

    # --- FICHA DE POSTULACION: solo para POSTULAR. REVISAR/DESCARTAR se
    #     dejan sin ficha, solo con la accion recomendada. Nada se elimina.
    def _ficha_fila(fila):
        if fila["decision_sugerida"] == "POSTULAR":
            return pd.Series(generar_ficha_postulacion(fila))
        if fila["decision_sugerida"] == "REVISAR":
            return pd.Series(("", "", "", "Revisar manualmente"))
        return pd.Series(("", "", "", "No priorizar"))

    nuevas[["porque_conviene", "encaje_perfil", "mensaje_sugerido", "accion_recomendada"]] = (
        nuevas.apply(_ficha_fila, axis=1)
    )
    cant_postular = conteo_decision.get("POSTULAR", 0)

    # --- categoria detectada: solo informativa, no toca decision/prioridad
    nuevas["categoria_detectada"] = nuevas.apply(detectar_categoria, axis=1)

    # --- top 5: se calcula ACA (antes del Excel) porque tambien lo usa
    #     la hoja RESUMEN, ademas del mensaje final de mas abajo.
    #     'nuevas' ya viene ordenada (decision_sugerida, prioridad,
    #     relevancia), asi que el top 5 es directamente head(5).
    top5_lineas = []
    top5_items = []
    for i, (_, fila) in enumerate(nuevas.head(5).iterrows(), start=1):
        top5_lineas.append(
            f"{i}. {fila['titulo']} - {fila['empresa']} - "
            f"{fila['fuente']} - {fila['decision_sugerida']} - {fila['categoria_detectada']}"
        )
        top5_items.append({
            "n": i, "titulo": fila["titulo"], "empresa": fila["empresa"],
            "fuente": fila["fuente"], "decision": fila["decision_sugerida"],
            "categoria": fila["categoria_detectada"],
        })
    top5_texto = "\n".join(top5_lineas) if top5_lineas else "(sin ofertas)"

    fecha = datetime.date.today().isoformat()
    # fecha+hora en el nombre del archivo del dia: si corro el script varias
    # veces en el mismo dia, cada corrida genera su propio Excel en vez de
    # pisar el anterior. 'fecha' (solo dia) se mantiene igual para el
    # historial acumulado (esa sigue siendo la marca de "primera vez visto").
    # El archivo va a resultados/YYYY-MM/; vistos.json e historial_trabajos.xlsx
    # se quedan en la carpeta principal (construir_ruta_excel no los toca).
    archivo = construir_ruta_excel(es_test=es_test)

    # --- bandeja de acciones: solo sugiere, no postula ni envia nada -----
    acciones_df, conteo_acciones = construir_acciones(nuevas)

    # --- postulacion asistida/automatica (respeta MODO_POSTULACION) ------
    # En modo test los links "nuevos" son en realidad avisos repetidos que
    # ya se vieron antes, asi que por default no se procesan para postular.
    # PROBAR_POSTULACION_EN_TEST=True permite probar el flujo completo
    # (apertura/chequeos/confirmacion) igual, sin esperar ofertas nuevas
    # reales -- sigue sin tocar vistos.json/historial_trabajos.xlsx (eso
    # ya esta fuera de este bloque) y cada resultado queda marcado como
    # prueba de TEST.
    if not es_test:
        acciones_df = ejecutar_postulaciones(acciones_df)
    elif PROBAR_POSTULACION_EN_TEST:
        acciones_df = ejecutar_postulaciones(acciones_df, es_test=True)
    else:
        acciones_df["estado_postulacion"] = ""
        acciones_df["resultado_postulacion"] = ""
        acciones_df["detalle_postulacion"] = ""

    datos_resumen = {
        "fecha_hora": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "modo_ejecucion": MODO_EJECUCION,
        "duracion": _formatear_duracion(inicio_ejecucion),
        "archivo": str(archivo),
        "cant_nuevos": cant_nuevos,
        "cant_repetidos": cant_repetidos,
        "postular": conteo_decision.get("POSTULAR", 0),
        "revisar": conteo_decision.get("REVISAR", 0),
        "descartar": conteo_decision.get("DESCARTAR", 0),
        "es_test": es_test,
        "top5": top5_items,
        "conteo_acciones": conteo_acciones,
    }

    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        crear_hoja_resumen(writer, datos_resumen)

        nuevas.to_excel(writer, sheet_name="TODOS", index=False)
        for fuente in nuevas["fuente"].unique():
            nuevas[nuevas["fuente"] == fuente].to_excel(writer, sheet_name=str(fuente)[:31], index=False)
        acciones_df.to_excel(writer, sheet_name="ACCIONES", index=False)

        aplicar_formato_excel(writer, nuevas, datos_resumen)

        # RESUMEN primero en el archivo, y sacar la hoja "Sheet" default
        # vacia que a veces deja openpyxl si nadie la uso.
        wb = writer.book
        if "RESUMEN" in wb.sheetnames:
            idx = wb.sheetnames.index("RESUMEN")
            if idx != 0:
                wb.move_sheet("RESUMEN", offset=-idx)
        if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"]["A1"].value is None:
            wb.remove(wb["Sheet"])

    # --- actualizar historial (vistos.json + historial_trabajos.xlsx) ---
    # En modo test NO se toca: 'nuevas' en ese caso son avisos repetidos,
    # ya estan en el historial, y el objetivo es justamente no contaminarlo.
    if not es_test:
        for link in nuevas["link"].fillna(""):
            if link:
                vistos[link] = fecha
        _guardar_vistos(vistos)
        _actualizar_historial_xlsx(nuevas, fecha)
    else:
        print("  MODO TEST: no se actualizó vistos.json ni historial_trabajos.xlsx.")

    print(f"\nListo: {archivo}")
    if es_test:
        print(f"  MODO TEST: {len(nuevas)} ofertas repetidas usadas para el Excel de prueba.")
    else:
        print(f"  {cant_nuevos} ofertas nuevas guardadas (guardadas: {cant_nuevos} de "
              f"{cant_nuevos + cant_repetidos} encontradas, {cant_repetidos} ya vistas)")
    print(f"  descartadas -> {descartadas_zona} por zona | "
          f"{filtradas} por filtro fino | {duplicadas} duplicadas")
    print("  Los links son clickeables (columna 'link' -> 'Abrir oferta').")
    # top5_texto ya se calculo mas arriba (se reusa aca para el mensaje final).

    if es_test:
        mensaje_final = (
            f"Buscador terminado en MODO TEST.\n\n"
            f"Archivo generado:\n{archivo}\n\n"
            f"Ofertas usadas para prueba: {len(nuevas)} ya vistas.\n"
            f"No se actualizó el historial.\n\n"
            f"POSTULAR: {conteo_decision.get('POSTULAR', 0)}\n"
            f"REVISAR: {conteo_decision.get('REVISAR', 0)}\n"
            f"DESCARTAR: {conteo_decision.get('DESCARTAR', 0)}\n\n"
            f"Top recomendadas:\n{top5_texto}"
        )
    else:
        mensaje_final = (
            f"Buscador terminado.\n\n"
            f"Archivo generado:\n{archivo}\n\n"
            f"Ofertas nuevas: {cant_nuevos}\n"
            f"Ya vistas antes: {cant_repetidos}\n\n"
            f"POSTULAR: {conteo_decision.get('POSTULAR', 0)}\n"
            f"REVISAR: {conteo_decision.get('REVISAR', 0)}\n"
            f"DESCARTAR: {conteo_decision.get('DESCARTAR', 0)}\n\n"
            f"Top recomendadas:\n{top5_texto}"
        )
    if cant_postular > 0:
        mensaje_final += (
            f"\n\nSe generaron fichas de postulación para {cant_postular} "
            f"ofertas marcadas como POSTULAR."
        )
    mensaje_final += (
        f"\n\nAcciones sugeridas:\n"
        f"POSTULAR HOY: {conteo_acciones.get('POSTULAR HOY', 0)}\n"
        f"REVISAR ANTES DE POSTULAR: {conteo_acciones.get('REVISAR ANTES DE POSTULAR', 0)}\n"
        f"REVISAR MANUALMENTE: {conteo_acciones.get('REVISAR MANUALMENTE', 0)}\n"
        f"NO ACCIONAR: {conteo_acciones.get('NO ACCIONAR', 0)}"
    )
    mensaje_final += f"\n\nDuración total: {_formatear_duracion(inicio_ejecucion)}."
    abrir_excel_si_corresponde(archivo)
    avisar_fin(mensaje_final)


if __name__ == "__main__":
    if not adquirir_lock():
        avisar_fin("El buscador ya se está ejecutando. Se canceló esta instancia.")
    else:
        try:
            main()
        finally:
            liberar_lock()

# ----------------------------------------------------------------------
# NOTAS
# ----------------------------------------------------------------------
# 1) Si Indeed y Bumeran siguen trayendo 0: Indeed bloquea por volumen
#    sin proxy, y la API interna de Bumeran pudo cambiar. Computrabajo
#    suele ser la mas estable.
# 2) FILTRO DE FECHA (3 dias): NO incluido. Computrabajo no devuelve
#    fecha confiable (viene vacia), asi que filtrar por fecha vaciaria
#    el Excel. Solo tendria sentido si Indeed vuelve a fun